# from hardcopy import bytestring_to_pdf
from rest_framework.permissions import IsAuthenticated
from rest_framework import mixins, viewsets, status
from rest_framework.response import Response
from django.contrib.gis.geos import Point, GEOSGeometry
from django.conf import settings
from django.utils import timezone
from ondoc.api.v1.common.serializers import SearchLeadSerializer
from django.utils.dateparse import parse_datetime
from weasyprint import HTML
from django.http import HttpResponse

from ondoc.api.v1.insurance.serializers import InsuranceCityEligibilitySerializer
from ondoc.api.v1.utils import html_to_pdf, generate_short_url
from ondoc.diagnostic.models import Lab
from ondoc.doctor.models import (Doctor, DoctorPracticeSpecialization, PracticeSpecialization, DoctorMobile, Qualification,
                                 Specialization, College, DoctorQualification, DoctorExperience, DoctorAward,
                                 DoctorClinicTiming, DoctorClinic, Hospital, SourceIdentifier, DoctorAssociation)

from ondoc.chat.models import ChatPrescription
from ondoc.insurance.models import InsuranceEligibleCities
from ondoc.lead.models import SearchLead
from ondoc.notification.models import EmailNotification
from ondoc.notification.rabbitmq_client import publish_message
# from ondoc.notification.sqs_client import publish_message
# from ondoc.notification.sqs_client import publish_message
from django.template.loader import render_to_string

from ondoc.procedure.models import IpdProcedure, IpdProcedureLead
from . import serializers
from ondoc.common.models import Cities, PaymentOptions, UserConfig, DeviceDetails, LastUsageTimestamp, AppointmentHistory
from ondoc.common.utils import send_email, send_sms
from ondoc.authentication.backends import JWTAuthentication, WhatsappAuthentication
from django.core.files.uploadedfile import SimpleUploadedFile, TemporaryUploadedFile, InMemoryUploadedFile
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
import random
import string
import base64
import logging
import datetime
import re
from django.db.models import Count
from io import BytesIO
import requests
from PIL import Image as Img
import os
import math
from decimal import Decimal

logger = logging.getLogger(__name__)

class CitiesViewSet(viewsets.GenericViewSet):

    def get_queryset(self):
        return Cities.objects.all().order_by('name')

    def list(self, request):
        filter_text = request.GET.get('filter', None)
        if not filter_text:
            response = [{'value': city.id, 'name': city.name} for city in self.get_queryset()]
        else:
            response = [{'value': city.id, 'name': city.name} for city in self.get_queryset().filter(name__istartswith=filter_text)]
        return Response(response)


class ServicesViewSet(viewsets.GenericViewSet):

    def generatepdf(self, request):
        from ondoc.api.v1.utils import generate_short_url
        content = None
        try:
            coded_data = request.data.get('content')
            if isinstance(coded_data, list):
                coded_data = coded_data[0]
            coded_data += "=="
            content = base64.b64decode(coded_data).decode()
        except Exception as e:
            logger.error("Error in decoding base64 content with exception - " + str(e))

        if not content:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={'Content is required.'})
        pdf_file = HTML(string=content).write_pdf()
        random_string = ''.join([random.choice(string.ascii_letters + string.digits) for n in range(12)])
        name = random_string + '.pdf'
        file = SimpleUploadedFile(name, pdf_file, content_type='application/pdf')
        chat = ChatPrescription.objects.create(name=name, file=file)
        prescription_url = "{}{}{}".format(settings.BASE_URL,
                                           "/api/v1/common/chat_prescription/",
                                           chat.name)
        short_url = generate_short_url(prescription_url)
        return Response({"url": short_url})

    def generate_pdf_template(self, request):
        from ondoc.api.v1.utils import generate_short_url
        context = {key: value for key, value in request.data.items()}
        if context.get('_updatedAt'):
            context['updated_at'] = parse_datetime(context.get('_updatedAt'))
        content = render_to_string("email/chat_prescription/body.html", context=context)
        random_string = ''.join([random.choice(string.ascii_letters + string.digits) for _ in range(12)])
        patient_profile = context.get('profile')
        patient_name = ''
        if patient_profile:
            patient_name = patient_profile.get('name', '')
        filename = 'dp_{}_{}_{}.pdf'.format('_'.join(patient_name.lower().split()), datetime.datetime.now().date(),
                                            random_string)
        file = html_to_pdf(content, filename)
        chat = ChatPrescription.objects.create(name=filename, file=file)
        prescription_url = "{}{}{}".format(settings.BASE_URL,
                                           "/api/v1/common/chat_prescription/",
                                           chat.name)
        short_url = generate_short_url(prescription_url)
        return Response({"url": short_url})

    def send_email(self, request):
        context = {key: value for key, value in request.data.items()}
        if context.get('_updatedAt'):
            context['updated_at'] = parse_datetime(context.get('_updatedAt'))
        serializer = serializers.EmailServiceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        to = serializer.validated_data.get('to')
        cc = serializer.validated_data.get('cc')
        to = list(set(to)) if isinstance(to, list) else []
        cc = list(set(cc)) if isinstance(cc, list) else []
        content = render_to_string("email/chat_prescription/body.html", context=context)
        subject = serializer.validated_data.get('subject')
        send_email(to, cc, subject, content)
        return Response({"status": "success"})

    def send_sms(self, request):
        serializer = serializers.SMSServiceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        text = serializer.validated_data.get('text')
        phone_number = serializer.validated_data.get('phone_number')
        phone_number = list(set(phone_number))
        send_sms(text, phone_number)
        return Response({"status": "success"})

    def download_pdf(self, request, name=None):
        chat_prescription = ChatPrescription.objects.filter(name=name).first()
        response = HttpResponse(chat_prescription.file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=%s' % chat_prescription.name
        return response


class SmsServiceViewSet(viewsets.GenericViewSet):
    #authentication_classes = (JWTAuthentication, )

    def send_sms(self, request):
        serializer = serializers.SMSServiceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        text = serializer.validated_data.get('text')
        phone_number = serializer.validated_data.get('phone_number')

        #phone_number = [request.user.phone_number] if request.user else []
        #phone_number = list(set(phone_number))
        send_sms(text, [phone_number])
        return Response({"status": "success"})


class UpdateXlsViewSet():
    fields = ['longitude', 'latitude', 'radius', 'specialty_id', 'lab_id', 'type', 'test_id']
    required_headers = fields+['result_count', 'url']

    def update(self, request):

        search_count_column = None
        serializer = serializers.XlsSerializer(data=request.FILES)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active

        rows = [row for row in sheet.rows]
        #columns = {i+1: column.value.strip().lower() for i, column in enumerate(rows[0]) if column.value}
        self.header = {column.value.strip().lower(): i + 1  for i, column in enumerate(rows[0]) if column.value}
        search_count_column = self.header.get('result_count')
        url_column = self.header.get('url')
        validation_url_column = self.header.get('validation_url')

        for i in range(2, len(rows)+1):
            data = self.get_data(i,sheet)
            output = self.get_result_count(data)
            if output!=0:
                sheet.cell(row=i, column=search_count_column).value = output[0]
                sheet.cell(row=i, column=url_column).value = output[1]
                sheet.cell(row=i, column=validation_url_column).value = output[2]

        response = HttpResponse(content=save_virtual_workbook(wb),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
        return response

    def get_data(self,row,sheet):
        data={}
        for key in self.fields:
            value = sheet.cell(row=row, column=self.header.get(key)).value
            if isinstance(value, str):
                value = value.strip().lower()
            if key in ['specialty_id','test_id']:
                value = [int(x) for x in str(value).split(',')] if value else []

            data[key] = value
        return data

    def get_result_count(self, data):
        if data['type']=='doctor' :
            return self.get_doctor_count(data);
        elif data['type']=='lab' :
            return self.get_lab_count(data);
        else :
            return 0

    def get_doctor_count(self, data):
        specialty_id = data['specialty_id']
        latitude = data['latitude']
        longitude = data['longitude']
        max_distance = 1000*data['radius']

        results = Doctor.objects.all()
        results = (Doctor.objects.filter(
            doctorpracticespecializations__specialization__in=specialty_id
            ) if len(specialty_id)>0 else Doctor.objects.all())

        search_count = results.filter(
            hospitals__location__distance_lte=(Point(longitude, latitude), max_distance),
            is_live=True,
            is_test_doctor=False,
            is_internal=False,
            hospitals__is_live=True).distinct().count()
        url = "{domain}/opd/searchresults?min_fees=0&max_fees=1500&sort_on=distance&is_available=false&is_female=false&doctor_name=&hospital_name=&conditions=&specializations={specialty_id}&lat={latitude}&long={longitude}&force_location=true".format(domain=settings.CONSUMER_APP_DOMAIN, specialty_id=','.join([str(x) for x in specialty_id]),latitude=str(latitude), longitude=str(longitude))
        validation_url = url+"&max_distance={max_distance}".format(max_distance=max_distance/1000)
        url = url + "&max_distance=20"
        return (search_count, url, validation_url)

    def get_lab_count(self, data):
        test_id = data['test_id']
        #lab_id = data['lab_id']
        latitude = data['latitude']
        longitude = data['longitude']
        max_distance = 1000*data['radius']

        filter = {'location__distance_lte': (Point(longitude, latitude), max_distance)}
        if len(test_id)>0:
            filter.update({
                'lab_pricing_group__available_lab_tests__test_id__in': test_id,
                'lab_pricing_group__available_lab_tests__enabled': True
            })
        count = 0

        search = Lab.objects.filter(is_test_lab=False, is_live=True,
                                          lab_pricing_group__isnull=False).filter(**filter)

        if len(test_id)>0:
            count = search.annotate(count=Count('id')).filter(count__gte=len(test_id)).count()
        else :
            count = search.count()

        url = "{domain}/lab/searchresults?min_distance=0&min_price=0&max_price=20000&sort_on=distancel&lab_name=&test_ids={test_id}&lat={latitude}&long={longitude}&force_location=true".format(domain=settings.CONSUMER_APP_DOMAIN, test_id=','.join([str(x) for x in test_id]),latitude=str(latitude), longitude=str(longitude))
        validation_url = url+"&max_distance={max_distance}".format(max_distance=max_distance/1000)
        url = url + "&max_distance=20"
        return (count, url, validation_url)

class UpdateXlsViewSet1():

    def update(self, request):
        search_count_column = None
        serializer = serializers.XlsSerializer(data=request.FILES)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        columns = {i+1: column.value.strip().lower() for i, column in enumerate(rows[0]) if column.value}
        for key in columns.keys():
            if columns.get(key) == 'number_of_results_on_search_page':
                search_count_column = key
        if search_count_column:
            for i in range(2, len(rows)+1):
                sheet.cell(row=i, column=search_count_column).value = self.get_result_count(i, columns, sheet)
            response = HttpResponse(content=save_virtual_workbook(wb),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
            return response
        return None

    def prepare_query(self, column, value, query):
        if column == 'test_id' and value is not None and value != '':
            query.update({
                'lab_pricing_group__available_lab_tests__test_id__in': [int(value.strip()) for value in
                                                                        str(value).split(",")],
                'lab_pricing_group__available_lab_tests__enabled': True
            })
        elif column == 'lab_id' and value is not None and value != '':
            query.update({
                'id': int(value)
            })
        return query

    def get_result_count(self, row, columns, sheet):
        longitude, latitude, max_distance, search_count_column, report_type = None, None, None, None, None
        query = {}
        for key in columns.keys():
            cell_value = str(sheet.cell(row=row, column=key).value).strip() if sheet.cell(row=row,
                                                                                          column=key).value else None
            if columns.get(key) == 'longitude':
                longitude = float(cell_value) if cell_value else None
            elif columns.get(key) == 'latitude':
                latitude = float(cell_value) if cell_value else None
            elif columns.get(key) == 'radius_in_km':
                max_distance = int(cell_value) * 1000 if cell_value else None
            elif columns.get(key) == 'specialization_ids':
                specialization_ids = [int(value.strip()) for value in cell_value.split(",")] if cell_value else []
            elif columns.get(key) == 'type':
                report_type = cell_value
            self.prepare_query(columns.get(key), sheet.cell(row=row, column=key).value, query)

        if report_type == 'doctor':
            results = (Doctor.objects.filter(
                doctorpracticespecializations__specialization__in=specialization_ids
            ) if specialization_ids else Doctor.objects.all())
            search_count = results.filter(
                hospitals__location__distance_lte=(Point(longitude, latitude), max_distance),
                is_live=True,
                is_test_doctor=False,
                is_internal=False,
                hospitals__is_live=True).distinct().count()
        else:
            query.update({
                'location__distance_lte': (Point(longitude, latitude), max_distance)
            })
            search_count = Lab.objects.filter(is_test_lab=False, is_live=True,
                                              lab_pricing_group__isnull=False).filter(**query).distinct().count()
        return search_count


class DocViewset(viewsets.GenericViewSet):

    def get_doctor(self, doctor_identifier):
        si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.DOCTOR, unique_identifier=doctor_identifier).first()
        doctor_obj = None
        if si_obj:
            doctor_obj = Doctor.objects.get(pk=si_obj.reference_id)
        return doctor_obj

    def clean_data(self, value):
        if value and isinstance(value, str):
            return value.strip()
        return value


class UploadDoctorViewSet(DocViewset):

    def upload(self, request):
        serializer = serializers.DoctorXLSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        source = validated_data.get('source')
        batch = validated_data.get('batch')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}

        for i in range(2, len(rows) + 1):
            data = self.get_data(row=i, sheet=sheet, headers=headers)
            doctor = self.create_doctor(data, source, batch)
            self.map_doctor_specialization(doctor, data.get('practice_specialization'))
            self.add_doctor_phone_numbers(doctor, data.get('numbers'))

        # response = HttpResponse(content=save_virtual_workbook(wb),
        #                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
        return Response(data={'message': 'success'})

    def get_data(self, row, sheet, headers):
        gender_mapping = {value[1]: value[0] for value in Doctor.GENDER_CHOICES}

        gender = gender_mapping.get(self.clean_data(sheet.cell(row=row, column=headers.get('gender')).value))
        identifier = self.clean_data(sheet.cell(row=row, column=headers.get('identifier')).value)
        name = self.clean_data(sheet.cell(row=row, column=headers.get('doctor_name')).value)
        license = self.clean_data(sheet.cell(row=row, column=headers.get('license')).value)
        city = self.clean_data(sheet.cell(row=row, column=headers.get('city')).value)
        practicing_since = self.clean_data(sheet.cell(row=row, column=headers.get('practicing_since')).value)

        if practicing_since:
            try:
                practicing_since = int(practicing_since)
            except:
                print('Invalid Practicing since='+str(practicing_since))
                practicing_since = None



        practice_specialization_id = self.clean_data(sheet.cell(row=row, column=headers.get('practice_specialization_id')).value)
        practice_specialization = None
        if practice_specialization_id:
            practice_specialization = PracticeSpecialization.objects.filter(pk=practice_specialization_id).first()

        number_entry = []
        primary_number = self.clean_data(sheet.cell(row=row, column=headers.get('primary_number')).value)
        alternate_number_1 = self.clean_data(sheet.cell(row=row, column=headers.get('alternate_number_1')).value)
        alternate_number_2 = self.clean_data(sheet.cell(row=row, column=headers.get('alternate_number_2')).value)

        num = self.get_number(primary_number, True, city)
        if num:
            number_entry.append(num)

        num = self.get_number(alternate_number_1, False, city)
        if num:
            number_entry.append(num)

        num = self.get_number(alternate_number_2, False, city)
        if num:
            number_entry.append(num)

        image_url = self.clean_data(sheet.cell(row=row, column=headers.get('image_url')).value)

        data = {}
        if not license:
            license=''

        data['gender'] = gender
        data['identifier'] = identifier
        data['name'] = name
        data['practicing_since'] = practicing_since
        data['practice_specialization'] = practice_specialization
        data['numbers'] = number_entry
        data['image_url'] = image_url
        data['license'] = license
        return data


    def get_number(self, number, is_primary, city):
        code=[11,44,22,129,40,120,33,215,124]
        data = {}
        std_code = None
        phone = None
        number = str(number).lstrip('0').strip()
        number = re.sub('[^0-9]+', ' ', number).strip()
        comps = number.split(' ')
        #print(number)
        if not number:
            return None

        if len(comps)==3 and comps[0] and comps[1] and comps[2]:
            return {'std_code':comps[0],'number':comps[1]+comps[2],'is_primary':False}
        elif len(comps)==2 and comps[0] and comps[1]:
            return {'std_code':comps[0],'number':comps[1],'is_primary':False}
        elif len(comps)>3:
            print('invalid number' + str(number))

        for cd in code:
            if number.startswith(str(cd)):
                data['std_code'] = cd
                data['number'] = number.replace(str(cd), '', 1)
                data['is_primary'] = False
                return data
        #print(number)
        try:
            number = int(number)
            if number < 5000000000 or number > 9999999999:
                print('invalid number' + str(number))
                return None
        except Exception as e:
            print(e)
            print('invalid number while parsing '+str(number))
            return None

        return {'number': number, 'is_primary': is_primary}

    def create_doctor(self, data, source, batch):

        doctor = self.get_doctor(data.get('identifier'))
        if doctor:
            return doctor

        doctor = Doctor.objects.create(name=data['name'], license=data.get('license',''), gender=data['gender'],
                                                       practicing_since=data['practicing_since'], source=source, batch=batch)
        SourceIdentifier.objects.create(type=SourceIdentifier.DOCTOR, unique_identifier=data.get('identifier'), reference_id=doctor.id)
        self.save_image(data.get('image_url'),data.get('identifier'))
        return doctor

    def save_image(self, url, identifier):
        if url and identifier:
            path = settings.MEDIA_ROOT+'/temp/image/'+identifier + '.jpg'
            final_path = settings.MEDIA_ROOT+'/temp/final/'+identifier + '.jpg'
            if os.path.exists(path):
                return None

            r = requests.get(url)
            content = BytesIO(r.content)
            # if os.path.exists(path):
            #     os.remove(path)
            # if os.path.exists(final_path):
            #     os.remove(final_path)

            of = open(path, 'xb')
            of.write(content.read())
            of.close()
            img = Img.open(path)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            size = img.size
            bottom = math.floor(size[1]*.8)
            img = img.crop((0,0,size[0],bottom))

            img.save(final_path, format='JPEG')
            # new_image_io.tell()
            # ff = open(final_path, 'xb')
            # ff.write(new_image_io.read())
            # ff.close()

    def map_doctor_specialization(self, doctor, practice_specialization):
        if practice_specialization and doctor:
            DoctorPracticeSpecialization.objects.get_or_create(doctor=doctor, specialization=practice_specialization)

    def add_doctor_phone_numbers(self, doctor, numbers):
        for num in numbers:
            #print(num)
            DoctorMobile.objects.get_or_create(doctor=doctor, std_code=num.get('std_code'),number=num.get('number'), is_primary=num.get('is_primary'))

    def clean_data(self, value):
        if value and isinstance(value, str):
            return value.strip()
        return value


class UploadQualificationViewSet(DocViewset):

    def upload(self, request):
        serializer = serializers.XlsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, len(rows) + 1):
            data = self.get_data(i, sheet, headers)
            #doctor = self.get_doctor(i, sheet, headers)
            # qualification = self.get_qualification(i, sheet, headers)
            # specialization = self.get_specialization(i, sheet, headers)
            # college = self.get_college(i, sheet, headers)
            # passing_year = self.clean_data(sheet.cell(row=i, column=headers.get('passing_year')).value)
            DoctorQualification.objects.get_or_create(doctor=data.get('doctor'),
                                                      qualification=data.get('qualification'),
                                                      college=data.get('college'),
                                                      specialization=data.get('specialization'),
                                                      passing_year=data.get('passing_year'))

        return Response(data={'message': 'success'})

    def get_data(self, row, sheet, headers):
        identifier = self.clean_data(sheet.cell(row=row, column=headers.get('identifier')).value)
        doctor = self.get_doctor(identifier)

        qualification_id = self.clean_data(
            sheet.cell(row=row, column=headers.get('qualification_id')).value)
        qualification_name = self.clean_data(sheet.cell(row=row, column=headers.get('qualification')).value)
        qualification = self.get_qualification(qualification_id, qualification_name)

        specialization_id = self.clean_data(
            sheet.cell(row=row, column=headers.get('specialization_id')).value)
        specialization_name = self.clean_data(sheet.cell(row=row, column=headers.get('specialization')).value)
        specialization = self.get_specialization(specialization_id, specialization_name)

        college_id = self.clean_data(
            sheet.cell(row=row, column=headers.get('college_id')).value)
        college_name = self.clean_data(sheet.cell(row=row, column=headers.get('college')).value)
        college = self.get_college(college_id, college_name)

        passing_year = self.clean_data(sheet.cell(row=row, column=headers.get('passing_year')).value)

        data = {}
        data['doctor'] = doctor
        data['qualification'] = qualification
        data['specialization'] = specialization
        data['college'] = college
        data['passing_year'] = passing_year

        return data;


    def get_qualification(self, qualification_id, qualification_name):

        qualification = None
        if qualification_id:
            qualification = Qualification.objects.filter(pk=qualification_id).first()
        if not qualification:
            qualification = Qualification.objects.filter(name__iexact=qualification_name).first()
        if not qualification:
            qualification, create = Qualification.objects.get_or_create(name=qualification_name)

        return qualification


    def get_specialization(self, specialization_id, specialization_name):

        specialization = None
        if specialization_id:
            specialization = Specialization.objects.filter(pk=specialization_id).first()
        if not specialization:
            specialization = Specialization.objects.filter(name__iexact=specialization_name).first()
        if not specialization:
            specialization, create = Specialization.objects.get_or_create(name=specialization_name)

        return specialization

    def get_college(self, college_id, college_name):

        college = None
        if college_id:
            college = College.objects.filter(pk=college_id).first()
        if not college:
            college = College.objects.filter(name__iexact=college_name).first()
        if not college:
            college, create = College.objects.get_or_create(name=college_name)
        return college

class UploadExperienceViewSet(DocViewset):

    def upload(self, request):
        serializer = serializers.XlsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, len(rows) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor = self.get_doctor(identifier)
            hospital = self.clean_data(sheet.cell(row=i, column=headers.get('hospital')).value)
            if hospital:
                start_year = self.clean_data(sheet.cell(row=i, column=headers.get('start_year')).value)
                end_year = self.clean_data(sheet.cell(row=i, column=headers.get('end_year')).value)
                DoctorExperience.objects.get_or_create(doctor=doctor, start_year=start_year, end_year=end_year,
                                                   hospital=hospital)
        return Response(data={'message': 'success'})


class UploadMembershipViewSet(DocViewset):

    def upload(self, request):
        serializer = serializers.XlsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, len(rows) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor = self.get_doctor(identifier)
            member = self.clean_data(sheet.cell(row=i, column=headers.get('memberships')).value)

            if doctor and member:
                DoctorAssociation.objects.get_or_create(doctor=doctor, name=member)
        return Response(data={'message': 'success'})


class UploadAwardViewSet(DocViewset):

    def upload(self, request):
        serializer = serializers.XlsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, len(rows) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor = self.get_doctor(identifier)
            award = self.clean_data(sheet.cell(row=i, column=headers.get('award')).value)
            year = self.clean_data(sheet.cell(row=i, column=headers.get('year')).value)
            if award and year:
                DoctorAward.objects.get_or_create(doctor=doctor, year=year)
        return Response(data={'message': 'success'})


class UploadHospitalViewSet(DocViewset):

    def upload(self, request):
        serializer = serializers.XlsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data
        reverse_day_map = {value[1]: value[0] for value in DoctorClinicTiming.SHORT_DAY_CHOICES}

        file = validated_data.get('file')
        wb = load_workbook(file)
        sheet = wb.active
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}

        doctor_obj_dict = dict()
        hospital_obj_dict = dict()
        doc_clinic_obj_dict = dict()
        for i in range(2, len(rows) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor_obj = self.get_doctor(identifier)
            if not doctor_obj:
                print('Doctor not found for identifier: '+identifier)
                continue

            hospital_obj = self.get_hospital(i, sheet, headers, hospital_obj_dict)
            doc_clinic_obj = self.get_doc_clinic(doctor_obj, hospital_obj, doc_clinic_obj_dict)
            day_list = self.parse_day_range(sheet.cell(row=i, column=headers.get('day_range')).value, reverse_day_map)
            start, end = self.parse_timing(sheet.cell(row=i, column=headers.get('timing')).value)
            clinic_time_data = list()
            fees = self.clean_data(sheet.cell(row=i, column=headers.get('fee')).value)
            try:
                fees = int(fees)
            except Exception as e:
                print('invalid fees' + str(fees))
                fees = None

            for day in day_list:
                if fees:
                    temp_data = {
                        "doctor_clinic": doc_clinic_obj,
                        "day": day,
                        "start": start,
                        "end": end,
                        "fees": fees
                    }
                    clinic_time_data.append(DoctorClinicTiming(**temp_data))
            if clinic_time_data:
                DoctorClinicTiming.objects.bulk_create(clinic_time_data)

        return Response(data={'message': 'success'})

    def get_hospital(self, row, sheet, headers, hospital_obj_dict):
        hospital_identifier = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_url')).value)
        hospital_id = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_id')).value)

        hospital = None
        if hospital_id:
            hospital = Hospital.objects.filter(pk=hospital_id).first()
        if not hospital:
            si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.HOSPITAL,
                                                     unique_identifier=hospital_identifier).first()
            if si_obj:
                hospital = Hospital.objects.filter(pk=si_obj.reference_id).first()

        if not hospital:
            hospital_name = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_name')).value)
            address = self.clean_data(sheet.cell(row=row, column=headers.get('address')).value)
            location = self.parse_gaddress(self.clean_data(sheet.cell(row=row, column=headers.get('gaddress')).value))
            hospital = Hospital.objects.create(name=hospital_name, building=address, location=location)
            SourceIdentifier.objects.create(reference_id=hospital.id, unique_identifier=hospital_identifier,
                                                type=SourceIdentifier.HOSPITAL)

        return hospital

    def get_doc_clinic(self, doctor_obj, hospital_obj, doc_clinic_obj_dict):
        # print(doctor_obj, hospital_obj, "hello")
        # if doc_clinic_obj_dict.get((doctor_obj, hospital_obj)):
        #     doc_clinic_obj = doc_clinic_obj_dict.get((doctor_obj, hospital_obj))
        # else:
        doc_clinic_obj, is_field_created = DoctorClinic.objects.get_or_create(doctor=doctor_obj, hospital=hospital_obj)
        # doc_clinic_obj_dict[(doctor_obj, hospital_obj)] = doc_clinic_obj

        return doc_clinic_obj

    def parse_gaddress(self, address):
        pnt = None

        if address:
            address = address.strip().split("/")
            lat_long = address[-1]
            lat_long_list = lat_long.strip().split(",")
            if len(lat_long_list) == 2 and "null" not in lat_long_list:
                point_string = 'POINT(' + str(lat_long_list[1].strip()) + ' ' + str(lat_long_list[0].strip()) + ')'
                pnt = GEOSGeometry(point_string, srid=4326)
        return pnt

    def parse_day_range(self, day_range, reverse_day_map):
        temp_list = day_range.split(',')
        days_list = list()
        for dr in temp_list:
            dr = dr.strip()
            rng_str = dr.split("-")
            if len(rng_str) == 1:
                day = rng_str[0].strip()
                if reverse_day_map.get(day) is None:
                    print('invalid day ' + str(day))
                else:
                    days_list.append(reverse_day_map[day])
            elif len(rng_str) == 2:
                s = rng_str[0].strip()
                e = rng_str[1].strip()
                if reverse_day_map.get(s) is None or reverse_day_map.get(e) is None:
                    print('invalid day range ' + str(day_range))
                else:
                    start = reverse_day_map[s]
                    end = reverse_day_map[e]
                    counter = start

                    while True:
                        val = counter % 7
                        days_list.append(val)
                        if val == end:
                            break
                        counter += 1

            else:
                print('invalid day string ' + day_range)
        return days_list

    def parse_timing(self, timing):
        tlist = timing.strip().split("-")
        start = None
        end = None
        if len(tlist) == 2:
            start = self.time_to_float(tlist[0])
            end = self.time_to_float(tlist[1])
        if start >= end:
            print('Invalid time string ' + timing)

        return start, end

    def time_to_float(self, time):
        hour_min, am_pm = time.strip().split(" ")
        hour, minute = hour_min.strip().split(":")
        hour = self.hour_to_int(int(hour.strip()), am_pm)
        minute = self.min_to_float(int(minute.strip()))
        return hour + minute

    def hour_to_int(self, hour, am_pm):
        if am_pm == "PM":
            if hour < 12:
                hour += 12
        elif am_pm == "AM":
            if hour >= 12:
                hour = 0
        return hour

    def min_to_float(self, minute):
        NUM_SLOTS_MIN = 2
        # print(float(minute)/60, "HI", minute, "HELLO", round(minute/60,2))
        minute = round(round(float(minute) / 60, 2) * NUM_SLOTS_MIN) / NUM_SLOTS_MIN
        return minute

    # def get_doctor(self, doctor_identifier):
    #     si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.DOCTOR, unique_identifier=doctor_identifier).first()
    #     doctor_obj = None
    #     if si_obj:
    #         doctor_obj = Doctor.objects.get(pk=si_obj.reference_id)
    #     return doctor_obj

    # def get_doctor(self, row, sheet, headers, doctor_obj_dict):
    #     doctor_identifier = self.clean_data(sheet.cell(row=row, column=headers.get('doctor_identifier')).value)
    #     if doctor_obj_dict.get(doctor_identifier):
    #         doctor_obj = doctor_obj_dict.get(doctor_identifier)
    #     else:
    #         # doctor_id = self.clean_data(sheet.cell(row=row, column=headers.get('doctor_id')).value)
    #         si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.DOCTOR, unique_identifier=doctor_identifier).first()
    #         doctor_obj = None
    #         if si_obj:
    #             doctor_obj = Doctor.objects.get(pk=si_obj.reference_id)
    #         doctor_obj_dict[doctor_identifier] = doctor_obj
    #     return doctor_obj

    # def clean_data(self, value):
    #     if value and isinstance(value, str):
    #         return value.strip()
    #     return value


class SearchLeadViewSet(viewsets.GenericViewSet):

    def create(self, request):
        serializer = SearchLeadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        html_body = ''
        for k, v in serializer.validated_data.items():
            html_body += '{} : {}\n'.format(k, v)
        for email in settings.OPS_EMAIL_ID:
            EmailNotification.publish_ops_email(email, html_body, 'New Search Lead')
        return Response({'msg': 'success'}, status=status.HTTP_200_OK)


class GetPaymentOptionsViewSet(viewsets.GenericViewSet):
    def get_queryset(self):
        return None

    def return_queryset(self, request):
        params = request.query_params
        from_app = params.get("from_app", False)
        if from_app:
            queryset = PaymentOptions.objects.filter(is_enabled=True).order_by('-priority')
        else:
            queryset = PaymentOptions.objects.filter(is_enabled=True).order_by('-priority')
        return queryset

    def list(self, request):
        queryset = self.return_queryset(request)
        options = PaymentOptions.build_payment_option(queryset)

        return Response(options)


class GetSearchUrlViewSet(viewsets.GenericViewSet):

    def search_url(self, request):
        params = request.query_params
        specialization_ids = params.get("specialization", '')
        from_app = params.get("from_app", False)
        test_ids = params.get("test", '')
        lat = params.get("lat", 28.4485)  # if no lat long then default to gurgaon
        long = params.get("long", 77.0759)

        if from_app == False:

            opd_search_url = "%s/opd/searchresults?specializations=%s" \
                             "&lat=%s&long=%s" \
                             % (settings.BASE_URL, specialization_ids, lat, long)
            tiny_opd_search_url = generate_short_url(opd_search_url)

            lab_search_url = "%s/lab/searchresults?test_ids=%s" \
                             "&lat=%s&long=%s" \
                             % (settings.BASE_URL, test_ids, lat, long)
            tiny_lab_search_url = generate_short_url(lab_search_url)

            return Response({"opd_search_url": tiny_opd_search_url, "lab_search_url": tiny_lab_search_url})

        else:

            opd_search_url = "docprm://docprime.com/opd/searchresults?specializations=%s" \
                             "&lat=%s&long=%s" \
                             % (specialization_ids, lat, long)

            lab_search_url = "docprm://docprime.com/lab/searchresults?test_ids=%s" \
                             "&lat=%s&long=%s" \
                             % (test_ids, lat, long)

            return Response({"opd_search_url": opd_search_url, "lab_search_url": lab_search_url})



class GetKeyDataViewSet(viewsets.GenericViewSet):

    def list(self, request):

        parameters = request.query_params
        key = parameters.get('key')
        if not key:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        queryset = UserConfig.objects.filter(key__iexact=key)
        if not queryset:
            return Response(status=status.HTTP_404_NOT_FOUND)
        resp = []
        for info in queryset:
            data_key = info.key
            data_key = data_key.lower()
            data = info.data
            resp.append({'data': data, 'key': data_key})
        return Response(resp)


class AllUrlsViewset(viewsets.GenericViewSet):

    def list(self, request):
        parameters = request.query_params
        key = parameters.get('query')
        if not key:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        key = key.lower()
        if len(key)<3:
            return Response(dict())


        from ondoc.location.models import EntityUrls
        from ondoc.location.models import CompareSEOUrls
        e_urls = list(EntityUrls.objects.filter(url__startswith=key, is_valid=True).values_list('url', flat=True))[:5]
        c_urls = list(CompareSEOUrls.objects.filter(url__startswith=key).values_list('url', flat=True))[:5]
        result = e_urls + c_urls
        return Response(dict(enumerate(result)))


class DeviceDetailsSave(viewsets.GenericViewSet):

    def save(self, request):
        user = request.user if request.user and request.user.is_authenticated else None

        last_usage_serializer = serializers.LastUsageTimestampSerializer(data=request.data)
        last_usage_serializer.is_valid(raise_exception=True)
        last_usage_validated_data = last_usage_serializer.validated_data

        add_or_update_device_details = False
        add_or_update_usage_time = True
        device_validated_data = None
        if last_usage_validated_data['source'] == AppointmentHistory.DOC_APP:
            device_serializer = serializers.DeviceDetailsSerializer(data=request.data)
            device_serializer.is_valid(raise_exception=True)
            device_validated_data = device_serializer.validated_data
            if 'device_id' in device_validated_data and device_validated_data.get('device_id'):
                add_or_update_device_details = True
                if 'last_ping_time' in device_validated_data and len(device_validated_data) == 2:
                    add_or_update_usage_time = False

        try:
            if add_or_update_device_details and device_validated_data:
                device_details_queryset = DeviceDetails.objects.filter(device_id=device_validated_data.get('device_id'))
                device_details = device_details_queryset.first()
                if device_details:
                    device_details_queryset.update(**device_validated_data, user=user)
                else:
                    if not 'data' in device_validated_data:
                        device_validated_data['data'] = {}
                    device_details = DeviceDetails.objects.create(**device_validated_data, user=user)
                last_usage_validated_data['device_id'] = device_details.id

            if user and add_or_update_usage_time:
                last_usage_queryset = LastUsageTimestamp.objects.filter(phone_number=user.phone_number)
                last_usage_details = last_usage_queryset.first()
                last_usage_validated_data['phone_number'] = int(user.phone_number)
                last_usage_validated_data['last_app_open_timestamp'] = datetime.datetime.now()
                if last_usage_details:
                    last_usage_queryset.update(**last_usage_validated_data)
                else:
                    LastUsageTimestamp.objects.create(**last_usage_validated_data)
        except Exception as e:
            logger.error("Something went wrong while saving last_usage_timestmap and device details - " + str(e))
            return Response("Error adding last_usage_timestmap and device details - " + str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({"status": 1, "message": "last_usage_timestmap and device details added"}, status=status.HTTP_200_OK)


class AppointmentPrerequisiteViewSet(viewsets.GenericViewSet):

    authentication_classes = (JWTAuthentication,)
    permission_classes = (IsAuthenticated, )

    def pre_booking(self, request):
        user = request.user
        insurance = user.active_insurance
        if not insurance or (user.is_authenticated and hasattr(request,'agent')):
            return Response({'prescription_needed': False})

        serializer = serializers.AppointmentPrerequisiteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        valid_data = serializer.validated_data

        user_profile = valid_data.get('profile', None)
        if not user_profile.is_insured_profile:
            return Response({'prescription_needed': False})

        lab = valid_data.get('lab')
        lab_pricing_group = lab.lab_pricing_group
        available_lab_test_qs = lab_pricing_group.available_lab_tests.all().filter(test__in=valid_data.get('lab_test'))
        tests_amount = Decimal(0)
        for available_lab_test in available_lab_test_qs:
            if available_lab_test.test.is_package and insurance.insurance_plan.plan_usages.get('member_package_limit'):
                if user_profile.is_insurance_package_limit_exceed():
                    return Response({'prescription_needed': True})
            agreed_price = available_lab_test.custom_agreed_price if available_lab_test.custom_agreed_price else available_lab_test.computed_agreed_price
            tests_amount = tests_amount + agreed_price

        # start_date = valid_data.get('start_date').date()

        resp = insurance.validate_limit_usages(tests_amount)

        return Response(resp)


class SiteSettingsViewSet(viewsets.GenericViewSet):
    authentication_classes = (JWTAuthentication,)

    def get_settings(self, request):
        params = request.query_params

        lat = params.get('latitude', None)
        long = params.get('longitude', None)

        insurance_availability = False

        if request.user and request.user.is_authenticated and request.user.active_insurance:
            insurance_availability = True

        if lat and long and not insurance_availability:
            data = {
                'latitude': lat,
                'longitude': long
            }

            serializer = InsuranceCityEligibilitySerializer(data=data)
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data
            city_name = InsuranceEligibleCities.get_nearest_city(data.get('latitude'), data.get('longitude'))
            if city_name:
                insurance_availability = True

        settings = {
            'insurance_availability': insurance_availability
        }

        return Response(data=settings)


class DepartmentRouting(viewsets.GenericViewSet):
    authentication_classes = (WhatsappAuthentication,)

    def get_department(self, request):
        params = request.query_params
        phone_number = params.get('phone_number', None)
        department_id = settings.CHAT_SOT_DEPARTMENT_ID
        department_name = 'Whatsapp'

        try:
            phone_number = int(phone_number)

            ipd_lead_active = IpdProcedureLead.check_if_lead_active(phone_number, 30)

            if ipd_lead_active:
                department_id = settings.CHAT_IPD_DEPARTMENT_ID
                department_name = 'IPD'
            else:
                pass

            resp = {
                'department_id': department_id,
                'department_name': department_name
            }
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(resp)