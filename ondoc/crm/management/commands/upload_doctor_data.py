from django.core.management.base import BaseCommand
import concurrent.futures
from django.conf import settings
from io import BytesIO
from openpyxl import load_workbook
import requests
import re
from PIL import Image as Img
import os
import math
from django.core.files.storage import default_storage

from ondoc.crm.constants import constants
from ondoc.doctor.models import (Doctor, DoctorPracticeSpecialization, PracticeSpecialization, DoctorMobile, Qualification,
                                 Specialization, College, DoctorQualification, DoctorExperience, DoctorAward,
                                 DoctorClinicTiming, DoctorClinic, Hospital, SourceIdentifier, DoctorAssociation)
from django.contrib.gis.geos import Point, GEOSGeometry
from django.contrib.contenttypes.models import ContentType

from ondoc.authentication.models import SPOCDetails, QCModel
from django.db import transaction


class Command(BaseCommand):
    help = 'Upload doctors via Excel'

    def add_arguments(self, parser):
        parser.add_argument('source', type=str, help='data source')
        parser.add_argument('batch', type=int, help='data batch')
        parser.add_argument('url', type=str, help='data url')
        parser.add_argument('lines', type=int, help='number of excel lines')


    def handle(self, *args, **options):
        print(options)
        source = options['source']
        batch = options['batch']
        url = options['url']
        lines = options['lines']

        r = requests.get(url)
        content = BytesIO(r.content)
        wb = load_workbook(content)
        sheets = wb.worksheets

        error_log = []

        doctor = UploadDoctor(error_log)
        qualification = UploadQualification(error_log)
        experience = UploadExperience(error_log)
        membership = UploadMembership(error_log)
        award = UploadAward(error_log)
        hospital = UploadHospital(error_log)
        specialization = UploadSpecialization(error_log)

        #doctor.p_image(sheets[0], source, batch)


        with transaction.atomic():
            doctor.upload(sheets[0], source, batch, lines)
            qualification.upload(sheets[1], lines)
            experience.upload(sheets[2], lines)
            membership.upload(sheets[3], lines)
            award.upload(sheets[4], lines)
            hospital.upload(sheets[5], source, batch, lines)
            specialization.upload(sheets[6], lines)


class Doc():

    def __init__(self):
        if not hasattr(self, 'log_arr'):
            raise Exception('error log not initialized')

    def log_error(self, line_number, message):
        self.log_arr.append({'line number': line_number, 'message': message})

    def get_doctor(self, doctor_identifier):
        si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.DOCTOR, unique_identifier=doctor_identifier).first()
        doctor_obj = None
        if si_obj:
            doctor_obj = Doctor.objects.filter(pk=si_obj.reference_id).first()
        return doctor_obj

    def clean_data(self, value):
        if value and isinstance(value, str):
            return value.strip()
        return value


    def get_number(self, number, is_primary, city, source, line_no=0):
        code=[11,44,22,129,40,120,33,215,124]
        data = {}
        std_code = None
        phone = None
        number = str(number).lstrip('0').strip()
        number = re.sub('[^0-9]+', ' ', number).strip()
        comps = number.split(' ')
        #print(number)
        if not number:
            return None

        if source.lower()=='google':
            return self.get_google_number(number, comps, source, is_primary)

        if len(comps)==3 and comps[0] and comps[1] and comps[2]:
            return {'std_code': comps[0], 'number': comps[1] + comps[2], 'is_primary': False, 'source': source}
        elif len(comps)==2 and comps[0] and comps[1]:
            if len(comps[0])>4:
                return {'number': comps[0] + comps[1], 'is_primary': is_primary, 'source': source}
            else:
                return {'std_code': comps[0], 'number': comps[1], 'is_primary': False, 'source': source}
        elif len(comps)>3:
            self.log_error(line_no, ' Invalid number : {}'.format(number))
            print('invalid number' + str(number))

        for cd in code:
            if number.startswith(str(cd)):
                data['std_code'] = cd
                data['number'] = number.replace(str(cd), '', 1)
                data['is_primary'] = False
                data['source'] = source
                return data
        #print(number)
        try:
            number = int(number)
            if number < 5000000000 or number > 9999999999:
                self.log_error(line_no, ' Invalid number : {}'.format(number))
                print('invalid number' + str(number))
                return None
        except Exception as e:
            print(e)
            self.log_error(line_no, ' Invalid number while parsing  : {}'.format(number))
            print('invalid number while parsing '+str(number))
            return None

        return {'number': number, 'is_primary': is_primary, 'source': source}

    def get_google_number(self, number, comps, source, is_primary):
        if len(comps)==3 and comps[0] and comps[1] and comps[2]:
            return {'std_code': comps[0], 'number': comps[1] + comps[2], 'is_primary': False, 'source': source}
        elif len(comps)==2 and comps[0] and comps[1]:
            return {'number': comps[0] + comps[1], 'is_primary': is_primary, 'source': source}
        elif len(comps)>3:
            print('invalid number' + str(number))



def s_image(batch, url, identifier):
    if url and identifier:
        #path = settings.MEDIA_ROOT+'/temp/image/'+identifier + '.jpg'
        #final_path = settings.MEDIA_ROOT+'/temp/final/'+identifier + '.jpg'
        path = 'temp/image/'+identifier + '.jpg'
        final_path = 'temp/final/'+identifier + '.jpg'

        if default_storage.exists(path):
            return None
        # if os.path.exists(path):
        #     return None

        # file = default_storage.open('storage_test', 'w')
        # file.write('storage contents')
        # file.close()


        r = requests.get(url)
        content = BytesIO(r.content)

        # if os.path.exists(path):
        #     os.remove(path)
        # if os.path.exists(final_path):
        #     os.remove(final_path)

        # of = open(path, 'xb')
        # of.write(content.read())
        # of.close()

        file = default_storage.open(path, 'wb')
        file.write(content.read())
        file.close()

        # r = requests.get(url)
        # content = BytesIO(r.content)
        ff = default_storage.open(path, 'rb')
        img = Img.open(ff)
        if img.mode != 'RGB':
            img = img.convert('RGB')
        size = img.size
        bottom = math.floor(size[1]*.8)
        img = img.crop((0,0,size[0],bottom))

        buffer = BytesIO()
        img.save(buffer, format='JPEG')
        buffer.seek(0)
        #print(len(buffer))
        #buffer.tell()

        cropped_file = default_storage.open(final_path, 'wb')
        cropped_file.write(buffer.read())
        cropped_file.close()


class UploadDoctor(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'Doctor Details'
        super().__init__()

    def p_image(self, sheet, source, batch):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        counter = 0
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            images = []

            for i in range(2, len(rows) + 1):
                identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
                url = self.clean_data(sheet.cell(row=i, column=headers.get('url')).value)
                if url:
                    images.append({'url':url, 'identifier':identifier, 'batch':batch})

            future_to_images = {executor.submit(s_image, image['batch'],
                                                image['url'],
                                                image['identifier'] ): image for image in images}

            counter=0
            for future in concurrent.futures.as_completed(future_to_images):
                img = future_to_images[future]
                data = future.result()
                counter += 1
                print('counter= '+str(counter))

            # for i in range(2, len(rows) + 1):
            #     print('processing = '+str(i))
            #     identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            #     url = self.clean_data(sheet.cell(row=i, column=headers.get('url')).value)
            #     if url:
            #         counter +=1
            #         print('found count='+str(counter))
            #         #print('url not found')
            #         try:
            #             future = executor.submit(s_image, batch, url, identifier).result()

            #             #self.s_image(batch, url, identifier)
            #         except Exception as e:
            #             print('exception '+str(e))

    def upload(self, sheet, source, batch, lines, user=None):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}

        for i in range(2, min(len(rows), lines) + 1):
            data = self.get_data(row=i, sheet=sheet, headers=headers, user=user)

            try:
                doctor = self.create_doctor(data, source, batch, i)
            except Exception as e:
                self.log_error(i, 'Invalid data for doctor. ({})'.format(e))
                print('error' + str(e))

            # self.map_doctor_specialization(doctor, data.get('practice_specialization'))
            try:
                self.add_doctor_phone_numbers(doctor, data.get('numbers'))
            except Exception as e:
                self.log_error(i, 'Invalid data for phone number. ({})'.format(e))
                print('error' + str(e))

    def get_data(self, row, sheet, headers, user=None):
        gender_mapping = {value[1]: value[0] for value in Doctor.GENDER_CHOICES}
        gender = gender_mapping.get(self.clean_data(sheet.cell(row=row, column=headers.get('gender')).value))
        identifier = self.clean_data(sheet.cell(row=row, column=headers.get('identifier')).value)
        name = self.clean_data(sheet.cell(row=row, column=headers.get('doctor_name')).value)
        license = self.clean_data(sheet.cell(row=row, column=headers.get('license')).value)
        city = self.clean_data(sheet.cell(row=row, column=headers.get('city')).value)
        practicing_since = self.clean_data(sheet.cell(row=row, column=headers.get('practicing_since')).value)
        is_license_verified = self.clean_data(
            sheet.cell(row=row, column=headers.get('is_license_verified')).value)
        enabled = self.clean_data(sheet.cell(row=row, column=headers.get('enabled')).value)
        if user and (user.is_member_of(constants['SUPER_QC_GROUP']) or user.is_member_of(
                constants['QC_GROUP_NAME']) or user.is_superuser):
            onboarding_status = self.clean_data(sheet.cell(row=row, column=headers.get('onboarding_status')).value)
            data_status = self.clean_data(sheet.cell(row=row, column=headers.get('data_status')).value)
            enabled_for_online_booking = self.clean_data(
                sheet.cell(row=row, column=headers.get('enabled_for_online_booking')).value)
            is_live = self.clean_data(sheet.cell(row=row, column=headers.get('is_live')).value)
        else:
            onboarding_status = Doctor.NOT_ONBOARDED
            data_status = QCModel.IN_PROGRESS
            enabled_for_online_booking = False
            is_live = False

        matrix_lead_id = self.clean_data(sheet.cell(row=row, column=headers.get('matrix_lead_id')).value)

        if practicing_since:
            try:
                practicing_since = int(practicing_since)
            except Exception as e:
                self.log_error(row, 'Invalid Practicing since : {} ({})'.format(practicing_since, e))
                print('Invalid Practicing since='+str(practicing_since))
                practicing_since = None

        # practice_specialization_id = self.clean_data(sheet.cell(row=row, column=headers.get('practice_specialization_id')).value)
        # practice_specialization = None
        # if practice_specialization_id:
        #     practice_specialization = PracticeSpecialization.objects.filter(pk=practice_specialization_id).first()

        number_entry = []
        primary_number = self.clean_data(sheet.cell(row=row, column=headers.get('primary_number')).value)
        alternate_number_1 = self.clean_data(sheet.cell(row=row, column=headers.get('alternate_number_1')).value)
        alternate_number_2 = self.clean_data(sheet.cell(row=row, column=headers.get('alternate_number_2')).value)
        source = self.clean_data(sheet.cell(row=row, column=headers.get('phone_no_source')).value)

        num = self.get_number(primary_number, True, city, source, row)
        if num:
            number_entry.append(num)

        num = self.get_number(alternate_number_1, False, city, source, row)
        if num:
            number_entry.append(num)

        num = self.get_number(alternate_number_2, False, city, source, row)
        if num:
            number_entry.append(num)

        image_url = self.clean_data(sheet.cell(row=row, column=headers.get('image_url')).value)

        data = {}
        if not license:
            license=''

        data['gender'] = gender
        data['identifier'] = identifier
        data['name'] = name
        data['practicing_since'] = practicing_since
        # data['practice_specialization'] = practice_specialization
        data['numbers'] = number_entry
        data['image_url'] = image_url
        data['license'] = license
        data['is_license_verified'] = is_license_verified
        data['onboarding_status'] = onboarding_status
        data['data_status'] = data_status
        data['enabled'] = enabled
        data['enabled_for_online_booking'] = enabled_for_online_booking
        data['is_live'] = is_live
        data['matrix_lead_id'] = matrix_lead_id
        return data

    def create_doctor(self, data, source, batch, line_no):
        doctor = self.get_doctor(data.get('identifier'))
        if doctor:
            return doctor
        doctor = Doctor.objects.create(name=data['name'], license=data.get('license', ''), gender=data['gender'],
                                       practicing_since=data['practicing_since'], source=source, batch=batch,
                                       enabled=data.get('enabled', False),
                                       enabled_for_online_booking=data.get('enabled_for_online_booking', False),
                                       data_status=data.get('data_status', QCModel.IN_PROGRESS),
                                       is_live=data.get('is_live', False),
                                       onboarding_status=data.get('onboarding_status', Doctor.NOT_ONBOARDED),
                                       is_license_verified=data.get('is_license_verified', False),
                                       matrix_lead_id=data.get('matrix_lead_id', None)
                                       )

        SourceIdentifier.objects.get_or_create(unique_identifier=data.get('identifier'), reference_id=doctor.id,
                                               type=SourceIdentifier.DOCTOR)
        #self.save_image(batch,data.get('image_url'),data.get('identifier'))
        return doctor

    def save_image(self, batch, url, identifier):
        if url and identifier:
            #path = settings.MEDIA_ROOT+'/temp/image/'+identifier + '.jpg'
            #final_path = settings.MEDIA_ROOT+'/temp/final/'+identifier + '.jpg'
            path = 'temp/image/'+identifier + '.jpg'
            final_path = 'temp/final/'+identifier + '.jpg'

            if default_storage.exists(path):
                return None
            # if os.path.exists(path):
            #     return None

            # file = default_storage.open('storage_test', 'w')
            # file.write('storage contents')
            # file.close()


            r = requests.get(url)
            content = BytesIO(r.content)

            # if os.path.exists(path):
            #     os.remove(path)
            # if os.path.exists(final_path):
            #     os.remove(final_path)

            # of = open(path, 'xb')
            # of.write(content.read())
            # of.close()

            file = default_storage.open(path, 'wb')
            file.write(content.read())
            file.close()

            # r = requests.get(url)
            # content = BytesIO(r.content)
            ff = default_storage.open(path, 'rb')
            img = Img.open(ff)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            size = img.size
            bottom = math.floor(size[1]*.8)
            img = img.crop((0,0,size[0],bottom))

            buffer = BytesIO()
            img.save(buffer, format='JPEG')
            buffer.seek(0)
            #print(len(buffer))
            #buffer.tell()

            cropped_file = default_storage.open(final_path, 'wb')
            cropped_file.write(buffer.read())
            cropped_file.close()
            #file.close()

            # new_image_io.tell()
            # ff = open(final_path, 'xb')
            # ff.write(new_image_io.read())
            # ff.close()

    def map_doctor_specialization(self, doctor, practice_specialization):
        if practice_specialization and doctor:
            DoctorPracticeSpecialization.objects.get_or_create(doctor=doctor, specialization=practice_specialization)

    def add_doctor_phone_numbers(self, doctor, numbers):
        for num in numbers:
            #print(num)
            #DoctorMobile.objects.get_or_create(doctor=doctor, std_code=num.get('std_code'),number=num.get('number'), is_primary=num.get('is_primary'))
            DoctorMobile.objects.get_or_create(doctor=doctor, std_code=num.get('std_code'),number=num.get('number'), defaults={'is_primary' : num.get('is_primary'), 'source' : num.get('source')})

    def clean_data(self, value):
        if value and isinstance(value, str):
            return value.strip()
        return value


class UploadQualification(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'Qualifications'
        super().__init__()

    def upload(self, sheet, lines):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, min(len(rows), lines) + 1):
            data = self.get_data(i, sheet, headers)
            #doctor = self.get_doctor(i, sheet, headers)
            # qualification = self.get_qualification(i, sheet, headers)
            # specialization = self.get_specialization(i, sheet, headers)
            # college = self.get_college(i, sheet, headers)
            # passing_year = self.clean_data(sheet.cell(row=i, column=headers.get('passing_year')).value)
            if data.get('doctor'):
                try:
                    DoctorQualification.objects.get_or_create(doctor=data.get('doctor'),
                                                              qualification=data.get('qualification'),
                                                              college=data.get('college'),
                                                              specialization=data.get('specialization'),
                                                              passing_year=data.get('passing_year'))
                except Exception as e:
                    self.log_error(i, 'Error saving doctor qualification. ()'.format(e))
                    print('error saving doctor qualification')


    def get_data(self, row, sheet, headers):
        identifier = self.clean_data(sheet.cell(row=row, column=headers.get('identifier')).value)
        doctor = self.get_doctor(identifier)

        qualification_id = self.clean_data(
            sheet.cell(row=row, column=headers.get('qualification_id')).value)
        qualification_name = self.clean_data(sheet.cell(row=row, column=headers.get('qualification')).value)
        if qualification_name and isinstance(qualification_name, str):
            qualification_name = re.sub(r'\s+', ' ', qualification_name)
        qualification = self.get_qualification(qualification_id, qualification_name, row)

        specialization_id = self.clean_data(
            sheet.cell(row=row, column=headers.get('specialization_id')).value)
        specialization_name = self.clean_data(sheet.cell(row=row, column=headers.get('specialization')).value)
        specialization = self.get_specialization(specialization_id, specialization_name, row)

        college_id = self.clean_data(
            sheet.cell(row=row, column=headers.get('college_id')).value)
        college_name = self.clean_data(sheet.cell(row=row, column=headers.get('college')).value)
        college = self.get_college(college_id, college_name)

        passing_year = self.clean_data(sheet.cell(row=row, column=headers.get('passing_year')).value)

        data = {}
        data['doctor'] = doctor
        data['qualification'] = qualification
        data['specialization'] = specialization
        data['college'] = college
        data['passing_year'] = passing_year

        return data;


    def get_qualification(self, qualification_id, qualification_name, line_no=0):

        qualification = None
        if qualification_id:
            qualification = Qualification.objects.filter(pk=qualification_id).first()
        if not qualification:
            qualification = Qualification.objects.filter(name__iexact=qualification_name).first()
        if not qualification and qualification_name:
            try:
                qualification, create = Qualification.objects.get_or_create(name=qualification_name)
            except Exception as e:
                self.log_error(line_no, 'Error while creating qualification. ({})'.format(e))

        return qualification

    def get_specialization(self, specialization_id, specialization_name, line_no=0):

        specialization = None
        if specialization_id:
            specialization = Specialization.objects.filter(pk=specialization_id).first()
        if not specialization:
            specialization = Specialization.objects.filter(name__iexact=specialization_name).first()
        if not specialization and specialization_name:
            try:
                specialization, create = Specialization.objects.get_or_create(name=specialization_name)
            except Exception as e:
                self.log_error(line_no, 'Error while creating Specialization. ({})'.format(e))
        return specialization

    def get_college(self, college_id, college_name, line_no=0):

        college = None
        if college_id:
            college = College.objects.filter(pk=college_id).first()
        if not college:
            college = College.objects.filter(name__iexact=college_name).first()
        if not college and college_name:
            try:
                college, create = College.objects.get_or_create(name=college_name)
            except Exception as e:
                self.log_error(line_no, 'Error while creating college. ({})'.format(e))
        return college

class UploadExperience(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'Experience'
        super().__init__()

    def upload(self, sheet, lines):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, min(len(rows), lines) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor = self.get_doctor(identifier)
            hospital = self.clean_data(sheet.cell(row=i, column=headers.get('hospital')).value)
            if hospital:
                start_year = self.clean_data(sheet.cell(row=i, column=headers.get('start_year')).value)
                end_year = self.clean_data(sheet.cell(row=i, column=headers.get('end_year')).value)
                try:
                    DoctorExperience.objects.get_or_create(doctor=doctor, start_year=start_year, end_year=end_year,
                                                   hospital=hospital)
                except Exception as e:
                    self.log_error(i, 'Error while creating doctor experience. ({})'.format(e))
                    print('error' + str(e))


class UploadSpecialization(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'Specialization'
        super().__init__()

    def upload(self, sheet, lines):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, min(len(rows), lines) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            sp_id = self.clean_data(sheet.cell(row=i, column=headers.get('specialization_id')).value)
            doctor = self.get_doctor(identifier)
            practice_specialization = None
            if sp_id:
                try:
                    practice_specialization = PracticeSpecialization.objects.filter(pk=sp_id).first()
                    if practice_specialization and doctor:
                        try:
                            DoctorPracticeSpecialization.objects.get_or_create(doctor=doctor, specialization=practice_specialization)
                        except Exception as e:
                            self.log_error(i, 'Error while creating Doctor Practice Specialization. ({})'.format(e))
                            print('error' + str(e))
                except Exception as e2:
                    self.log_error(i, 'Error while creating Practice Specialization. ({})'.format(e))
                    print('error' + str(e2))


class UploadMembership(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'Membership'
        super().__init__()

    def upload(self, sheet, lines):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, min(len(rows), lines) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor = self.get_doctor(identifier)
            member = self.clean_data(sheet.cell(row=i, column=headers.get('memberships')).value)

            if doctor and member:
                try:
                    DoctorAssociation.objects.get_or_create(doctor=doctor, name=member)
                except Exception as e:
                    self.log_error(i, 'Error while creating doctor membership. ({})'.format(e))
                    print('error' + str(e))


class UploadAward(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'Awards'
        super().__init__()

    def upload(self, sheet, lines):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, min(len(rows), lines) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor = self.get_doctor(identifier)
            award = self.clean_data(sheet.cell(row=i, column=headers.get('award')).value)
            year = self.clean_data(sheet.cell(row=i, column=headers.get('year')).value)
            if award and year:
                try:
                    DoctorAward.objects.get_or_create(doctor=doctor, name=award, year=year)
                except Exception as e:
                    self.log_error(i, 'Error while creating doctor awards. ({})'.format(e))
                    print('error' + str(e))


class UploadHospital(Doc):

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'DoctorHospital'
        super().__init__()

    def upload(self, sheet, source, batch, lines):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        reverse_day_map = {value[1].lower(): value[0] for value in DoctorClinicTiming.SHORT_DAY_CHOICES}
        type_choices_mapping = {value[1]: value[0] for value in DoctorClinicTiming.TYPE_CHOICES}
        doctor_obj_dict = dict()
        hospital_obj_dict = dict()
        doc_clinic_obj_dict = dict()
        hospital_obj = None
        for i in range(2, min(len(rows), lines) + 1):
            identifier = self.clean_data(sheet.cell(row=i, column=headers.get('identifier')).value)
            doctor_obj = self.get_doctor(identifier)
            if not doctor_obj:
                if identifier:
                    self.log_error(i, 'Doctor not found for identifier : {}.'.format(identifier))
                    print('Doctor not found for identifier: '+identifier)
                else:
                    print('Doctor not found for identifier: ')
                continue

            hospital_obj = self.get_hospital(i, sheet, headers, hospital_obj_dict, source, batch)
            if not hospital_obj:
                self.log_error(i, 'Hospital not found')
                print('hospital not found')
                continue
            followup_duration = self.clean_data(sheet.cell(row=i, column=headers.get('followup_duration')).value)
            followup_charges = self.clean_data(sheet.cell(row=i, column=headers.get('followup_charges')).value)
            try:
                followup_duration = int(followup_duration)
            except Exception as e:
                self.log_error(i, 'invalid followup_duration : {} ({})'.format(followup_duration, e))
                print('invalid followup_duration' + str(followup_duration))
                followup_duration = 0
            try:
                followup_charges = int(followup_charges)
            except Exception as e:
                self.log_error(i, 'invalid followup_charges : {} ({})'.format(followup_charges, e))
                print('invalid followup_charges' + str(followup_charges))
                followup_charges = 0
            doc_clinic_obj = self.get_doc_clinic(doctor_obj, hospital_obj, doc_clinic_obj_dict, followup_duration, followup_charges)
            day_list = self.parse_day_range(sheet.cell(row=i, column=headers.get('day_range')).value, reverse_day_map)
            start, end = self.parse_timing(sheet.cell(row=i, column=headers.get('timing')).value)
            clinic_time_data = list()
            fees = self.clean_data(sheet.cell(row=i, column=headers.get('fee')).value)
            type = type_choices_mapping.get(self.clean_data(sheet.cell(row=i, column=headers.get('type')).value), None)
            deal_price = self.clean_data(sheet.cell(row=i, column=headers.get('deal_price')).value)
            mrp = self.clean_data(sheet.cell(row=i, column=headers.get('mrp')).value)

            if not type:
                self.log_error(i, 'Invalid type for clinic timing')
                raise Exception('Invalid type for clinic timing')

            try:
                fees = int(fees)
            except Exception as e:
                self.log_error(i, 'Invalid fees : {} ({})'.format(fees, e))
                print('invalid fees' + str(fees))
                fees = None
            try:
                deal_price = int(deal_price)
            except Exception as e:
                self.log_error(i, 'Invalid deal_price : {} ({})'.format(deal_price, e))
                print('invalid deal_price' + str(deal_price))
                deal_price = None
            try:
                mrp = int(mrp)
            except Exception as e:
                self.log_error(i, 'Invalid mrp : {} ({})'.format(mrp, e))
                print('invalid mrp' + str(mrp))
                mrp = None

            for day in day_list:
                if fees is not None and day is not None and start is not None and end is not None and start != end:
                    temp_data = {
                        "doctor_clinic": doc_clinic_obj,
                        "day": day,
                        "start": start,
                        "end": end,
                        "fees": fees,
                        "deal_price": deal_price,
                        "mrp": mrp,
                        "type": type
                    }
                    try:
                        DoctorClinicTiming.objects.get_or_create(**temp_data)
                    except:
                        self.log_error(i, 'Error while creating doctor clinic timing. ({})'.format(e))
                        print('query error')
                    #clinic_time_data.append(DoctorClinicTiming(**temp_data))
            # if clinic_time_data:
            #     DoctorClinicTiming.objects.bulk_create(clinic_time_data)

            primary_number = self.clean_data(sheet.cell(row=i, column=headers.get('clinic_contact_1')).value)
            alternate_number_1 = self.clean_data(sheet.cell(row=i, column=headers.get('clinic_contact_2')).value)
            alternate_number_2 = self.clean_data(sheet.cell(row=i, column=headers.get('clinic_contact_3')).value)
            ph_source = self.clean_data(sheet.cell(row=i, column=headers.get('phone_no_source')).value)


            number_entry = []
            if not ph_source:
                ph_source=''
            # else:
            #     print(ph_source)

            num = self.get_number(primary_number, True, '', ph_source)
            if num:
                number_entry.append(num)

            num = self.get_number(alternate_number_1, False, '', ph_source)
            if num:
                number_entry.append(num)

            num = self.get_number(alternate_number_2, False, '', ph_source)
            if num:
                number_entry.append(num)

            try:
                self.add_clinic_phone_numbers(hospital_obj, number_entry)
            except Exception as e:
                self.log_error(i, 'Error while creating doctor clinic phone number. ({})'.format(e))
                print(e)


    def add_clinic_phone_numbers(self, hospital, numbers):
        ct = ContentType.objects.get_for_model(hospital)
        for num in numbers:
            #print(num)
            SPOCDetails.objects.get_or_create(content_type=ct, object_id=hospital.id, std_code=num.get('std_code'),number=num.get('number'), defaults={'contact_type':1, 'source':num.get('source')})


    def get_hospital(self, row, sheet, headers, hospital_obj_dict, source, batch):
        hospital_identifier = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_unique_identifier')).value)
        hospital_id = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_id')).value)

        hospital = None
        if hospital_id:
            hospital = Hospital.objects.filter(pk=hospital_id).first()
        if not hospital:
            si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.HOSPITAL,
                                                     unique_identifier=hospital_identifier).first()
            if si_obj:
                hospital = Hospital.objects.filter(pk=si_obj.reference_id).first()
        try:
            if not hospital:
                hospital_name = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_name')).value)
                building = self.clean_data(sheet.cell(row=row, column=headers.get('building')).value)
                if not building:
                    building=''

                city = self.clean_data(sheet.cell(row=row, column=headers.get('city')).value)
                location = self.parse_gaddress(self.clean_data(sheet.cell(row=row, column=headers.get('hospital_lat_long')).value))
                hospital = Hospital.objects.create(name=hospital_name, building=building, city=city, country='India', location=location, source=source, batch=batch, enabled_for_online_booking=False)
                SourceIdentifier.objects.create(reference_id=hospital.id, unique_identifier=hospital_identifier,
                                                    type=SourceIdentifier.HOSPITAL)
        except Exception as e:
            self.log_error(row, 'Error while creating hospital. ({})'.format(e))
            print(str(e))

        return hospital

    def get_doc_clinic(self, doctor_obj, hospital_obj, doc_clinic_obj_dict, followup_duration, followup_charges):
        # print(doctor_obj, hospital_obj, "hello")
        # if doc_clinic_obj_dict.get((doctor_obj, hospital_obj)):
        #     doc_clinic_obj = doc_clinic_obj_dict.get((doctor_obj, hospital_obj))
        # else:
        doc_clinic_obj, is_field_created = DoctorClinic.objects.get_or_create(doctor=doctor_obj, hospital=hospital_obj,                                                                              
                                                                              defaults={
                                                                              'followup_charges':followup_charges,
                                                                              'followup_duration':followup_duration,
                                                                              'enabled_for_online_booking': True})
        # doc_clinic_obj_dict[(doctor_obj, hospital_obj)] = doc_clinic_obj

        return doc_clinic_obj

    def parse_gaddress(self, address):
        pnt = None

        if address:
            address = address.strip().split("/")
            lat_long = address[-1]
            lat_long_list = lat_long.strip().split(",")
            if len(lat_long_list) == 2 and "null" not in lat_long_list:
                point_string = 'POINT(' + str(lat_long_list[1].strip()) + ' ' + str(lat_long_list[0].strip()) + ')'
                pnt = GEOSGeometry(point_string, srid=4326)
        return pnt

    def parse_day_range(self, day_range, reverse_day_map):

        if not day_range:
            return list()
        temp_list = day_range.split(',')
        days_list = list()
        for dr in temp_list:
            dr = dr.strip()
            rng_str = dr.split("-")
            if len(rng_str) == 1:
                day = rng_str[0].strip().lower()
                if reverse_day_map.get(day) is None:
                    print('invalid day ' + str(day))
                else:
                    days_list.append(reverse_day_map[day])
            elif len(rng_str) == 2:
                s = rng_str[0].strip().lower()
                e = rng_str[1].strip().lower()
                if reverse_day_map.get(s) is None or reverse_day_map.get(e) is None:
                    print('invalid day range ' + str(day_range))
                else:
                    start = reverse_day_map[s]
                    end = reverse_day_map[e]
                    counter = start

                    while True:
                        val = counter % 7
                        days_list.append(val)
                        if val == end:
                            break
                        counter += 1

            else:
                print('invalid day string ' + day_range)
        return days_list

    def parse_timing(self, timing):

        if not timing:
            return None, None
        print(timing)
        tlist = timing.strip().split("-")
        start = None
        end = None
        if len(tlist) == 2:
            start = self.time_to_float(tlist[0])
            end = self.time_to_float(tlist[1])
        if start >= end:
            print('Invalid time string ' + timing)

        return start, end

    def time_to_float(self, time):
        hour_min, am_pm = time.strip().split()
        hour, minute = hour_min.strip().split(":")
        hour = self.hour_to_int(int(hour.strip()), am_pm)
        minute = self.min_to_float(int(minute.strip()))
        return hour + minute

    def hour_to_int(self, hour, am_pm):
        if am_pm == "PM":
            if hour < 12:
                hour += 12
        elif am_pm == "AM":
            if hour >= 12:
                hour = 0
        return hour

    def min_to_float(self, minute):
        NUM_SLOTS_MIN = 2
        # print(float(minute)/60, "HI", minute, "HELLO", round(minute/60,2))
        minute = round(round(float(minute) / 60, 2) * NUM_SLOTS_MIN) / NUM_SLOTS_MIN
        return minute

    # def get_doctor(self, doctor_identifier):
    #     si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.DOCTOR, unique_identifier=doctor_identifier).first()
    #     doctor_obj = None
    #     if si_obj:
    #         doctor_obj = Doctor.objects.get(pk=si_obj.reference_id)
    #     return doctor_obj

    # def get_doctor(self, row, sheet, headers, doctor_obj_dict):
    #     doctor_identifier = self.clean_data(sheet.cell(row=row, column=headers.get('doctor_identifier')).value)
    #     if doctor_obj_dict.get(doctor_identifier):
    #         doctor_obj = doctor_obj_dict.get(doctor_identifier)
    #     else:
    #         # doctor_id = self.clean_data(sheet.cell(row=row, column=headers.get('doctor_id')).value)
    #         si_obj = SourceIdentifier.objects.filter(type=SourceIdentifier.DOCTOR, unique_identifier=doctor_identifier).first()
    #         doctor_obj = None
    #         if si_obj:
    #             doctor_obj = Doctor.objects.get(pk=si_obj.reference_id)
    #         doctor_obj_dict[doctor_identifier] = doctor_obj
    #     return doctor_obj

    # def clean_data(self, value):
    #     if value and isinstance(value, str):
    #         return value.strip()
    #     return value
