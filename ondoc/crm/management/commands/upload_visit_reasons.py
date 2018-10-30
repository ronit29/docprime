from django.core.management.base import BaseCommand
from io import BytesIO
from openpyxl import load_workbook
import requests
from ondoc.doctor.models import (VisitReason, PracticeSpecialization, VisitReasonMapping)
import re


class Command(BaseCommand):
    help = 'Upload doctors via Excel'

    def add_arguments(self, parser):
        parser.add_argument('url', type=str, help='data url')

    def handle(self, *args, **options):
        print(options)
        url = options['url']
        r = requests.get(url)
        content = BytesIO(r.content)
        wb = load_workbook(content)
        # wb = load_workbook(url)
        sheets = wb.worksheets
        visit_reason = UploadVisitReason()
        visit_reason.upload(sheets[1], False)
        visit_reason.upload(sheets[2], True)


class UploadVisitReason:
    all_practice_specialization_ids = PracticeSpecialization.objects.all().values_list('pk', flat=True)

    def upload(self, sheet, is_primary):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, len(rows) + 1):
            datas = self.get_data(row=i, sheet=sheet, headers=headers, is_primary=is_primary)
            for data in datas:
                obj, created = VisitReasonMapping.objects.get_or_create(visit_reason_id=data.get('visit_reason_id'),
                                                                        practice_specialization_id=data.get(
                                                                            'practice_specialization_id'))
                if not created:
                    obj.is_primary = data.get('is_primary')
                    obj.save()

    def get_data(self, row, sheet, headers, is_primary):
        names = self.clean_data(sheet.cell(row=row, column=headers.get('name')).value)
        practice_specialization = self.clean_data(
            sheet.cell(row=row, column=headers.get('practice_specialization')).value)
        try:
            practice_specialization = int(practice_specialization)
        except:
            print('Invalid practice_specialization = ' + str(practice_specialization))
            return []

        if not practice_specialization in self.all_practice_specialization_ids:
            print('practice_specialization not found = ' + str(practice_specialization))
            return []

        all_name = names.split(',')
        all_names = []
        for name in all_name:
            name = name.strip()
            name = re.sub(r'\s+', ' ', name)
            name = name.capitalize()
            all_names.append(name)
        all_names = set(all_names)
        already_added = VisitReason.objects.all().values_list('name', flat=True)
        already_added = set(already_added)
        to_be_added = all_names - already_added
        all_visit_reasons_obj = []
        for name in to_be_added:
            all_visit_reasons_obj.append(VisitReason(name=name))
        VisitReason.objects.bulk_create(all_visit_reasons_obj)

        visit_ids = VisitReason.objects.filter(name__in=all_names).values_list('pk', flat=True)

        datas = []
        for visit_id in visit_ids:
            data = {'visit_reason_id': visit_id, 'practice_specialization_id': practice_specialization,
                    'is_primary': is_primary}
            datas.append(data)
        return datas

    @staticmethod
    def clean_data(value):
        if value and isinstance(value, str):
            return value.strip()
        return value
