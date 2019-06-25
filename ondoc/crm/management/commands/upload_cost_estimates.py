from django.core.management.base import BaseCommand
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError, transaction
import requests
from openpyxl import load_workbook
from io import BytesIO
from ondoc.doctor.models import Hospital
from ondoc.procedure.models import IpdProcedureCostEstimate, IpdProcedure, IpdCostEstimateRoomType, \
    IpdCostEstimateRoomTypeMapping


class Command(BaseCommand):
    help = 'Upload cost estimate via Excel'

    def add_arguments(self, parser):
        parser.add_argument('url', type=str, help='data url')

    @transaction.atomic
    def handle(self, *args, **options):
        print(options)
        url = options['url']
        r = requests.get(url)
        content = BytesIO(r.content)
        wb = load_workbook(content)
        sheets = wb.worksheets
        cost_estimate = UploadCostEstimate()

        cost_estimate.upload(sheets[0])


class UploadCostEstimate:

    def __init__(self, log_arr=None) -> None:
        self.log_arr = log_arr
        self.sheet = 'IPD Cost Estimate Details'
        super().__init__()

    def log_error(self, line_number, message):
        self.log_arr.append({'line number': line_number, 'message': message})

    def upload(self, sheet):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        for i in range(2, len(rows) + 1):
            data = self.get_data(row=i, sheet=sheet, headers=headers)
            try:
                ipd_cost_estimate = IpdProcedureCostEstimate.objects.filter(ipd_procedure=data.get('ipd_procedure'), hospital=data.get('hospital')).first()
                if not ipd_cost_estimate:
                    ipd_cost_estimate = IpdProcedureCostEstimate.objects.create(
                        ipd_procedure=data.get('ipd_procedure'),
                        hospital=data.get('hospital'),
                        stay_duration=data.get('stay_duration')
                    )
                else:
                    ipd_cost_estimate.stay_duration = data.get('stay_duration')
                    ipd_cost_estimate.save()
                ipd_cost_estimate_room_type_mapping = IpdCostEstimateRoomTypeMapping.objects.filter(cost_estimate=ipd_cost_estimate, room_type=data.get('room_type')).first()
                if not ipd_cost_estimate_room_type_mapping:
                    IpdCostEstimateRoomTypeMapping.objects.create(
                        cost_estimate=ipd_cost_estimate,
                        room_type=data.get('room_type'),
                        cost=data.get('cost')
                    )
                else:
                    ipd_cost_estimate_room_type_mapping.cost = data.get('cost')
                    ipd_cost_estimate_room_type_mapping.save()
            except IntegrityError as e:
                self.log_error(i, 'Error in file - .'.format(e))

    def get_data(self, row, sheet, headers):
        procedure_id = self.clean_data(sheet.cell(row=row, column=headers.get('procedure_id')).value)
        hospital_id = self.clean_data(sheet.cell(row=row, column=headers.get('hospital_id')).value)
        stay_duration = self.clean_data(sheet.cell(row=row, column=headers.get('stay_duration')).value)
        room_type_id = self.clean_data(sheet.cell(row=row, column=headers.get('room_type_id')).value)
        cost = self.clean_data(sheet.cell(row=row, column=headers.get('cost')).value)
        try:
            ipd_procedure = IpdProcedure.objects.get(pk=procedure_id)
        except ObjectDoesNotExist:
            ipd_procedure = None
            self.log_error(row, 'Ipd procedure with id {} not found.'.format(procedure_id))

        try:
            hospital = Hospital.objects.get(pk=hospital_id)
        except ObjectDoesNotExist:
            hospital = None
            self.log_error(row, 'Hospital with id {} not found.'.format(hospital_id))
            # print('Hospital with ID ' + str(hospital_id) + ' not available')

        try:
            room_type = IpdCostEstimateRoomType.objects.get(pk=room_type_id)
        except ObjectDoesNotExist:
            room_type = None
            self.log_error(row, 'Room type with id {} not found.'.format(room_type_id))
            # print('Room type with ID '+ room_type_id + ' not available')

        try:
            stay_duration = int(stay_duration)
        except ValueError:
            self.log_error(row, 'Stay duration {} is not valid.'.format(stay_duration))

        data = {'ipd_procedure': ipd_procedure, 'hospital': hospital, 'stay_duration': stay_duration,
                'room_type': room_type, 'cost': cost}
        return data

    @staticmethod
    def clean_data(value):
        if value and isinstance(value, str):
            return value.strip()
        return value
