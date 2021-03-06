from __future__ import absolute_import, unicode_literals

import copy
import datetime
from datetime import timedelta
import json
import math
import traceback
from collections import OrderedDict
from io import BytesIO

from django.contrib.contenttypes.models import ContentType
from django.db.models import Q

import pytz
from django.db import transaction
from django.forms import model_to_dict
from django.utils import timezone
from openpyxl import load_workbook

from ondoc.api.v1.utils import aware_time_zone, log_requests_on, pg_seamless_hash, get_pg_secret_client_key
from ondoc.authentication.models import UserNumberUpdate, UserProfileEmailUpdate, Address
from ondoc.common.models import AppointmentMaskNumber
from ondoc.matrix.mongo_models import MatrixLog
from ondoc.notification.labnotificationaction import LabNotificationAction
from ondoc.notification import models as notification_models
from celery import task
import logging
from django.conf import settings
import requests
from rest_framework import status
from django.utils.safestring import mark_safe
from ondoc.notification.models import NotificationAction
import random
import string
from ondoc.api.v1.utils import RawSql

logger = logging.getLogger(__name__)


@task
def send_lab_notifications_refactored(data, *args, **kwargs):
    from ondoc.diagnostic import models as lab_models
    from ondoc.communications.models import LabNotification
    appointment_id = data.get('appointment_id', None)
    is_valid_for_provider = data.get('is_valid_for_provider', False)
    instance = lab_models.LabAppointment.objects.filter(id=appointment_id).first()
    if not instance or not instance.user:
        return
    try:
        test_list = []
        instance = lab_models.LabAppointment.objects.filter(id=appointment_id).first()
        if not instance or not instance.user:
            return
        if instance.status == lab_models.LabAppointment.COMPLETED:
            instance.generate_invoice()
        counter = 1
        is_masking_done = False
        while counter < 3:
            if is_masking_done:
                break
            else:
                counter = counter + 1
                is_masking_done = generate_appointment_masknumber(
                    ({'type': 'LAB_APPOINTMENT', 'appointment': instance}))
        lab_notification = LabNotification(instance)
        context = lab_notification.get_context()
        if lab_notification.notification_type not in (NotificationAction.LAB_APPOINTMENT_BOOKED, NotificationAction.LAB_APPOINTMENT_RESCHEDULED_BY_LAB,
                                                      NotificationAction.LAB_APPOINTMENT_RESCHEDULED_BY_PATIENT):
            lab_notification.send(is_valid_for_provider)
        else:
            # notification_type = lab_notification.notification_type
            # receivers = lab_notification.get_receivers(is_valid_for_provider)
            # from ondoc.communications.models import SMSNotification
            # sms_notification = SMSNotification(notification_type, context)
            # sms_notification.send(receivers.get('sms_receivers', []))

            content_type = ContentType.objects.get_for_model(instance)
            mask_no_obj = AppointmentMaskNumber.objects.filter(object_id=instance.id, content_type_id=content_type.id)
            app_test_mapping = instance.test_mappings.all()
            est = pytz.timezone(settings.TIME_ZONE)
            time_slot_start = instance.time_slot_start.astimezone(est)
            for data in app_test_mapping:
                if data.test:
                    test_list.append({'name': data.test.name, 'reference_code': data.test.reference_code})
            context = {'id': instance.id, 'lab_name': instance.lab.name if instance.lab else '',
                                'Patient_name': instance.profile_detail.get('name'), 'Gender':instance.profile_detail.get('gender'),
                                'DOB': instance.profile_detail.get('dob'), 'pickup_address': instance.get_pickup_address() if instance.is_home_pickup else None,
                                'lab_address': instance.lab.get_lab_address(), 'time_slot': time_slot_start.strftime("%I:%M%p"),
                                'Date':time_slot_start.date(), 'mask_number': mask_no_obj[0].mask_number if mask_no_obj and mask_no_obj[0].mask_number
                        else instance.address.get('phone_number'), 'test_list': test_list, 'client_code': 'CH343' if instance and instance.lab and instance.lab.network and instance.lab.network.id == 195 else '', 'lab_network_id': instance.lab.network.id if instance and instance.lab and instance.lab.network else None}

            lab_notification.send(is_valid_for_provider, overrided_context=context, email_obj=instance)

    except Exception as e:
        logger.error(str(e))

@task
def send_ipd_procedure_lead_mail(data):
    obj_id = data.get('obj_id')
    send_email = data.get('send_email')
    from ondoc.communications.models import EMAILNotification
    from ondoc.procedure.models import IpdProcedureLead
    from ondoc.matrix.tasks import create_or_update_lead_on_matrix
    instance = IpdProcedureLead.objects.filter(id=obj_id).first()
    if not instance:
        return
    if instance.matrix_lead_id:
        return
    if not instance.is_valid:
        return
    try:
        if send_email:
            emails = settings.IPD_PROCEDURE_CONTACT_DETAILS
            user_and_email = [{'user': None, 'email': email} for email in emails]
            email_notification = EMAILNotification(notification_type=NotificationAction.IPD_PROCEDURE_MAIL,
                                                   context={'instance': instance})
            email_notification.send(user_and_email)

        create_or_update_lead_on_matrix.apply_async(
            ({'obj_type': instance.__class__.__name__, 'obj_id': instance.id},), countdown=5)

    except Exception as e:
        logger.error(str(e))



@task
def send_lab_notifications(appointment_id):
    from ondoc.diagnostic import models as lab_models
    instance = lab_models.LabAppointment.objects.filter(id=appointment_id).first()
    if not instance:
        return
    if not instance.user:
        return
    # lab_managers = lab_models.LabManager.objects.filter(lab=instance.lab)
    lab_managers = instance.get_lab_admins()
    if instance.status == lab_models.LabAppointment.COMPLETED:
        LabNotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.LAB_INVOICE,
        )
        return
    if instance.status == lab_models.LabAppointment.ACCEPTED:
        LabNotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.LAB_APPOINTMENT_ACCEPTED,
        )
        return
    if instance.status == lab_models.LabAppointment.RESCHEDULED_PATIENT:
        for lab_manager in lab_managers:
            LabNotificationAction.send_to_lab_managers(
                instance, lab_manager, notification_models.NotificationAction.LAB_APPOINTMENT_RESCHEDULED_BY_PATIENT)
        LabNotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.LAB_APPOINTMENT_RESCHEDULED_BY_PATIENT,
        )
        return
    if instance.status == lab_models.LabAppointment.RESCHEDULED_LAB:
        LabNotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.LAB_APPOINTMENT_RESCHEDULED_BY_LAB,
        )
        return
    if instance.status == lab_models.LabAppointment.CANCELLED:
        for lab_manager in lab_managers:
            LabNotificationAction.send_to_lab_managers(
                instance, lab_manager, notification_models.NotificationAction.LAB_APPOINTMENT_CANCELLED)
        LabNotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.LAB_APPOINTMENT_CANCELLED,
        )
        return
    if instance.status == lab_models.LabAppointment.BOOKED:
        for lab_manager in lab_managers:
            LabNotificationAction.send_to_lab_managers(
                instance, lab_manager, notification_models.NotificationAction.LAB_APPOINTMENT_BOOKED)
        LabNotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.LAB_APPOINTMENT_BOOKED,
        )
        return


@task()
def send_opd_notifications_refactored(data, notification_type=None):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    appointment_id = data.get('appointment_id', None)
    is_valid_for_provider = data.get('is_valid_for_provider', False)
    notification_type = data.get('notification_type', None)
    instance = OpdAppointment.objects.filter(id=appointment_id).first()
    try:
        if not instance or not instance.user:
            return
        if instance.status == OpdAppointment.COMPLETED:
            instance.generate_invoice()
        if instance.status == OpdAppointment.ACCEPTED and instance.is_credit_letter_required_for_appointment() and not instance.is_payment_type_cod():
            instance.generate_credit_letter()
        counter = 1
        is_masking_done = False
        while counter < 3:
            if is_masking_done:
                break
            else:
                counter = counter + 1
                is_masking_done = generate_appointment_masknumber(
                    ({'type': 'OPD_APPOINTMENT', 'appointment': instance}))
    except Exception as e:
        logger.error(str(e))
    opd_notification = OpdNotification(instance, notification_type)
    opd_notification.send(is_valid_for_provider)


@task
def send_opd_notifications(appointment_id):
    from ondoc.authentication.models import GenericAdmin
    from ondoc.doctor.models import OpdAppointment
    instance = OpdAppointment.objects.filter(id=appointment_id).first()
    if not instance:
        return
    if not instance.user:
        return
    doctor_admins = GenericAdmin.get_appointment_admins(instance)
    if instance.user and instance.status == OpdAppointment.ACCEPTED:
        notification_models.NotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.APPOINTMENT_ACCEPTED,
        )
    elif instance.status == OpdAppointment.RESCHEDULED_PATIENT:
        for admin in doctor_admins:
            notification_models.NotificationAction.trigger(
                instance=instance,
                user=admin,
                notification_type=notification_models.NotificationAction.APPOINTMENT_RESCHEDULED_BY_PATIENT)
        if not instance.user:
            return
        notification_models.NotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.APPOINTMENT_RESCHEDULED_BY_PATIENT)
    elif instance.status == OpdAppointment.RESCHEDULED_DOCTOR:
        if not instance.user:
            return
        notification_models.NotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.APPOINTMENT_RESCHEDULED_BY_DOCTOR)
    elif instance.status == OpdAppointment.BOOKED:
        for admin in doctor_admins:
            notification_models.NotificationAction.trigger(
                instance=instance,
                user=admin,
                notification_type=notification_models.NotificationAction.APPOINTMENT_BOOKED)
        notification_models.NotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.APPOINTMENT_BOOKED)
    elif instance.status == OpdAppointment.CANCELLED:
        for admin in doctor_admins:
            notification_models.NotificationAction.trigger(
                instance=instance,
                user=admin,
                notification_type=notification_models.NotificationAction.APPOINTMENT_CANCELLED)
        if not instance.user:
            return
        notification_models.NotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.APPOINTMENT_CANCELLED)
    elif instance.status == OpdAppointment.COMPLETED:
        if not instance.user:
            return
        notification_models.NotificationAction.trigger(
            instance=instance,
            user=instance.user,
            notification_type=notification_models.NotificationAction.DOCTOR_INVOICE,
        )

@task(max_retries=1)
def send_opd_rating_message(appointment_id, type):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    from django.conf import settings
    data = {}
    name = ''
    try:
        if type == 'opd':
            appointment = OpdAppointment.objects.filter(id=appointment_id, is_rated=False, status=OpdAppointment.COMPLETED).first()
            name = appointment.doctor.name if appointment else None
        else:
            appointment = LabAppointment.objects.filter(id=appointment_id, is_rated=False, status=LabAppointment.COMPLETED).first()
            name = appointment.lab.name if appointment else None
        if appointment:
            number = appointment.user.phone_number
            data['phone_number'] = number
            app_url = settings.CONSUMER_APP_DOMAIN
            text_url = str(app_url)+ "/" + str(type) + "/appointment/" + str(appointment_id)
            text = '''You have successfully completed your appointment with %s . Rate your experience %s''' % (name, text_url)
            data['text'] = mark_safe(text)
            notification_models.SmsNotification.send_rating_link(data)
    except Exception as e:
        logger.error(str(e))
        pass

@task(bind=True, max_retries=5)
def set_order_dummy_transaction(self, order_id, user_id):
    from ondoc.account.models import Order, DummyTransactions
    from ondoc.insurance.models import UserInsurance
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.account.models import User
    from ondoc.account.mongo_models import PgLogs
    from ondoc.account.models import MerchantPayoutLog
    from ondoc.account.models import PgTransaction
    from ondoc.plus.models import PlusUser

    req_data = dict()
    try:
        order_row = Order.objects.filter(id=order_id).first()
        user = User.objects.filter(id=user_id).first()

        if not order_row or not user:
            raise Exception('order and user are required')
        if order_row.getTransactions():
            return
        if not order_row.dummy_transaction_allowed():
            raise Exception("Cannot create dummy payout for a child order.")

        appointment = order_row.getAppointment()
        if not appointment:
            raise Exception("No Appointment/UserPlanMapping/PlusUser found.")

        total_price = order_row.get_total_price()

        token = settings.PG_DUMMY_TRANSACTION_TOKEN
        headers = {
            "auth": token,
            "Content-Type": "application/json"
        }
        url = settings.PG_DUMMY_TRANSACTION_URL
        # insurance_data = order_row.get_insurance_data_for_pg()
        additional_data = order_row.get_additional_data_for_pg()

        if appointment.__class__.__name__ in ['LabAppointment', 'OpdAppointment']:
            if appointment.insurance and not order_row.is_parent() and not additional_data:
                MerchantPayoutLog.create_log(None, "refOrderId, insurerCode and refOrderNo not found for order id {}".format(order_row.id))
                raise Exception("refOrderId, insurerCode, refOrderNo details not found for order id {}".format(order_row.id))

        name = ''
        if isinstance(appointment, OpdAppointment) or isinstance(appointment, LabAppointment):
            name = appointment.profile.name

        if isinstance(appointment, UserInsurance) or isinstance(appointment, PlusUser):
            name = appointment.user.full_name

        req_data = {
            "customerId": user_id,
            "mobile": user.phone_number,
            "email": user.email or "dummyemail@docprime.com",
            "productId": order_row.product_id,
            "orderId": order_id,
            "name": name,
            "txAmount": 0,
            "couponCode": "",
            "couponAmt": str(total_price),
            "paymentMode": "DC",
            "AppointmentId": appointment.id,
            "buCallbackSuccessUrl": "",
            "buCallbackFailureUrl": ""
        }

        # req_data.update(insurance_data)
        req_data.update(additional_data)
        for key in req_data:
            req_data[key] = str(req_data[key])

        secret_key, client_key = get_pg_secret_client_key(order_row)
        filtered_pgdata = {k: v for k, v in req_data.items() if v is not None and v != ''}
        req_data.clear()
        req_data.update(filtered_pgdata)
        req_data['hash'] = PgTransaction.create_pg_hash(req_data, secret_key, client_key)

        response = requests.post(url, data=json.dumps(req_data), headers=headers)
        if response.status_code == status.HTTP_200_OK:
            if settings.SAVE_LOGS:
                save_pg_response.apply_async((PgLogs.DUMMY_TXN, order_id, None, response.json(), req_data, user_id,),
                                             eta=timezone.localtime(), queue=settings.RABBITMQ_LOGS_QUEUE)
            resp_data = response.json()
            #logger.error(resp_data)
            if resp_data.get("ok") is not None and resp_data.get("ok") == 1:
                tx_data = {}
                tx_data['user'] = user
                tx_data['product_id'] = order_row.product_id
                tx_data['order_no'] = resp_data.get('orderNo')
                tx_data['order_id'] = order_row.id
                tx_data['reference_id'] = appointment.id
                tx_data['type'] = DummyTransactions.CREDIT
                tx_data['amount'] = total_price
                tx_data['payment_mode'] = "DC"

                DummyTransactions.objects.create(**tx_data)
                #print("SAVED DUMMY TRANSACTION")
        else:
            raise Exception("Retry on invalid Http response status - " + str(response.content))

    except Exception as e:
        logger.error("Error in Setting Dummy Transaction of user with data - " + json.dumps(req_data) + " with exception - " + str(e))
        self.retry([order_id, user_id], countdown=300)


@task(bind=True, max_retries=5)
def activate_chat_plans(self, plus_user_id):
    from ondoc.plus.models import PlusPlans
    from ondoc.plus.models import PlusUser
    try:
        obj = PlusUser.objects.filter(id=plus_user_id).first()
        if not obj:
            raise Exception('No Plus User found with id - ' + str(plus_user_id))
        plan = obj.plan
        request_data = {}
        if not plan:
            raise Exception('Chat plan - Not able to find Plan')
        if plan.is_chat_included:
            request_data['mobileNo'] = obj.user.phone_number
            request_data['priorityType'] = PlusPlans.NORMAL_CHAT_PLAN if not obj.plan.chat_plans else obj.plan.chat_plans
            url = settings.CHAT_GOLD_API_URL + "addPriorityNumbers"
            auth_token = settings.CHAT_LAB_REPORT_API_TOKEN
            response = requests.post(url, data=json.dumps(request_data), headers={'Authorization': auth_token, 'Content-Type': 'application/json'})
            res = response.json()

    except Exception as e:
        logger.error("Error in Setting Chat Plan for Gold - " + " with exception - " + str(e))
        self.retry(plus_user_id, countdown=300)


# @task(bind=True, max_retries=5)
# def set_order_dummy_transaction_for_corporate(self, order_id, user_id):
#     from ondoc.account.models import PgTransaction
#     from ondoc.account.models import Order, DummyTransactions
#     from ondoc.plus.models import PlusUser
#     from ondoc.account.models import User
#     from ondoc.account.mongo_models import PgLogs
#     from ondoc.account.models import MerchantPayoutLog
#     req_data = dict()
#     try:
#         order_row = Order.objects.filter(id=order_id).first()
#         user = User.objects.filter(id=user_id).first()
#
#         if not order_row or not user:
#             raise Exception('order and user are required')
#         if order_row.getTransactions():
#             return
#         if not order_row.dummy_transaction_allowed():
#             raise Exception("Cannot create dummy payout for a child order.")
#
#         # appointment = order_row.getAppointment()
#         # if not appointment:
#         #     raise Exception("No Appointment/UserPlanMapping found.")
#
#         # total_price = order_row.get_total_price()
#         total_price = 0
#         token = settings.PG_DUMMY_TRANSACTION_TOKEN
#         headers = {
#             "auth": token,
#             "Content-Type": "application/json"
#         }
#         url = settings.PG_DUMMY_TRANSACTION_URL
#         action_data = order_row.action_data
#         if not action_data:
#             raise Exception("Can not create dummy transaction for empty action data")
#         profile = action_data.get('profile_detail')
#         if not profile:
#             raise Exception("Can not create dummy transaction for empty profile")
#         name = profile.get('name', '')
#         plus_user = PlusUser.objects.filter(user_id=user_id).order_by('-id').first()
#         # insurance_data = order_row.get_insurance_data_for_pg()
#
#         # if appointment.__class__.__name__ in ['LabAppointment', 'OpdAppointment']:
#         #     if appointment.insurance and not order_row.is_parent() and not insurance_data:
#         #         MerchantPayoutLog.create_log(None, "refOrderId, insurerCode and refOrderNo not found for order id {}".format(order_row.id))
#         #         raise Exception("refOrderId, insurerCode, refOrderNo details not found for order id {}".format(order_row.id))
#         #
#         # name = ''
#         # if isinstance(appointment, OpdAppointment) or isinstance(appointment, LabAppointment):
#         #     name = appointment.profile.name
#         #
#         # if isinstance(appointment, UserInsurance):
#         #     name = appointment.user.full_name
#
#         req_data = {
#             "customerId": user_id,
#             "mobile": user.phone_number,
#             "email": user.email or "dummyemail@docprime.com",
#             "productId": order_row.product_id,
#             "orderId": order_id,
#             "name": name,
#             "txAmount": 0,
#             "couponCode": "",
#             "couponAmt": str(total_price),
#             "paymentMode": "DC",
#             "AppointmentId": plus_user.id,
#             "buCallbackSuccessUrl": "",
#             "buCallbackFailureUrl": ""
#         }
#
#         # req_data.update(insurance_data)
#
#         for key in req_data:
#             req_data[key] = str(req_data[key])
#
#         secret_key, client_key = get_pg_secret_client_key(order_row)
#         filtered_pgdata = {k: v for k, v in req_data.items() if v is not None and v != ''}
#         req_data.clear()
#         req_data.update(filtered_pgdata)
#         req_data['hash'] = PgTransaction.create_pg_hash(req_data, secret_key, client_key)
#
#         response = requests.post(url, data=json.dumps(req_data), headers=headers)
#         save_pg_response.apply_async((PgLogs.DUMMY_TXN, order_id, None, response.json(), req_data, user_id,), eta=timezone.localtime(), queue=settings.RABBITMQ_LOGS_QUEUE)
#         if response.status_code == status.HTTP_200_OK:
#             resp_data = response.json()
#             #logger.error(resp_data)
#             if resp_data.get("ok") is not None and resp_data.get("ok") == 1:
#                 tx_data = {}
#                 tx_data['user'] = user
#                 tx_data['product_id'] = order_row.product_id
#                 tx_data['order_no'] = resp_data.get('orderNo')
#                 tx_data['order_id'] = order_row.id
#                 tx_data['reference_id'] = plus_user.id
#                 tx_data['type'] = DummyTransactions.CREDIT
#                 tx_data['amount'] = total_price
#                 tx_data['payment_mode'] = "DC"
#
#                 DummyTransactions.objects.create(**tx_data)
#                 #print("SAVED DUMMY TRANSACTION")
#         else:
#             raise Exception("Retry on invalid Http response status - " + str(response.content))
#
#     except Exception as e:
#         logger.error("Error in Setting Dummy Transaction of user with data - " + json.dumps(req_data) + " with exception - " + str(e))
#         self.retry([order_id, user_id], countdown=300)

@task
def send_offline_appointment_message(**kwargs):
    from ondoc.doctor.models import OfflineOPDAppointments
    from ondoc.communications.models import OfflineOpdAppointments
    appointment_id = kwargs.get('appointment_id')
    notification_type = kwargs.get('notification_type')
    receivers = kwargs.get('receivers')
    try:
        if appointment_id:
            appointment = OfflineOPDAppointments.objects.filter(id=appointment_id).first()
            offline_opd_appointment_comm = OfflineOpdAppointments(appointment=appointment,
                                                                  notification_type=notification_type,
                                                                  receivers=receivers)
            offline_opd_appointment_comm.send()
    except Exception as e:
        logger.error("Error sending " + str(type) + " message - " + str(e))

def generate_appointment_masknumber(data):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    # appointment_type = data.get('type')
    is_masking_done = False
    # try:
    #     is_mask_number = True
    #     is_maskable = True
    #     appointment = data.get('appointment', None)
    #     if not appointment:
    #         # logger.error("[CELERY ERROR: Incorrect values provided.]")
    #         raise Exception("Appointment not found, could not get mask number")
    #
    #     if appointment_type == 'OPD_APPOINTMENT':
    #         is_network_enabled_hospital = appointment.hospital
    #         if is_network_enabled_hospital:
    #             is_maskable = is_network_enabled_hospital.is_mask_number_required
    #         else:
    #             is_maskable = True
    #     elif data.get('type') == 'LAB_APPOINTMENT':
    #         is_network_enabled_lab = appointment.lab.network
    #         if is_network_enabled_lab:
    #             is_maskable = is_network_enabled_lab.is_mask_number_required
    #         else:
    #             is_maskable = True
    #
    #     phone_number = appointment.user.phone_number
    #     time_slot = appointment.time_slot_start
    #     updated_time_slot = time_slot + datetime.timedelta(days=1)
    #     validity_up_to = int((time_slot + datetime.timedelta(days=1)).timestamp())
    #     if not phone_number:
    #         raise Exception("phone Number could not found against appointment id - " + str(appointment.id))
    #     if is_maskable:
    #         request_data = {
    #             "ExpirationDate": validity_up_to,
    #             "FromId": appointment.id,
    #             "ToNumber": phone_number
    #         }
    #         url = settings.MATRIX_NUMBER_MASKING
    #         matrix_api_token = settings.MATRIX_API_TOKEN
    #         response = requests.post(url, data=json.dumps(request_data), headers={'Authorization': matrix_api_token,
    #                                                                               'Content-Type': 'application/json'})
    #
    #         if response.status_code != status.HTTP_200_OK or not response.ok:
    #             logger.info("[ERROR] Appointment could not be get Mask Number")
    #             logger.info("[ERROR] %s", response)
    #             # countdown_time = (2 ** self.request.retries) * 60 * 10
    #             # logging.error("Appointment sync with the Matrix System failed with response - " + str(response.content))
    #             # print(countdown_time)
    #             # self.retry([data], countdown=countdown_time)
    #
    #         mask_number = response.json()
    #     else:
    #         is_mask_number = False
    #         mask_number = phone_number
    #     if mask_number:
    #         existing_mask_number_obj = appointment.mask_number.filter(is_deleted=False).first()
    #         if existing_mask_number_obj:
    #             existing_mask_number_obj.is_deleted = True
    #             existing_mask_number_obj.save()
    #             AppointmentMaskNumber(content_object=appointment, mask_number=mask_number,
    #                                                  validity_up_to=updated_time_slot, is_mask_number=is_mask_number,
    #                                                  is_deleted=False).save()
    #             is_masking_done = True
    #         else:
    #             AppointmentMaskNumber(content_object=appointment, mask_number=mask_number,
    #                                                  validity_up_to=updated_time_slot, is_mask_number=is_mask_number,
    #                                                  is_deleted=False).save()
    #             is_masking_done = True
    #     else:
    #         raise Exception("Failed to generate Mask Number for the appointment - " + str(appointment.id))
    # except Exception as e:
    #     logger.error("Error in Celery. Failed to get mask number for appointment " + str(e))
    return is_masking_done

@task
def send_rating_update_message(number, text):
    data = {}
    data['phone_number'] = number
    data['text'] = mark_safe(text)
    try:
        notification_models.SmsNotification.send_rating_link(data)
    except Exception as e:
        logger.error("Error sending rating update sms")

@task
def send_appointment_reminder_message(number, patient_name, doctor, hospital_name, date):
    data = {}
    data['phone_number'] = number
    text = '''Dear %s, you have an appointment scheduled with %s at %s on %s''' % (patient_name, doctor, hospital_name, date)
    data['text'] = mark_safe(text)
    try:
        notification_models.SmsNotification.send_rating_link(data)
    except Exception as e:
        logger.error("Error sending Reminder message - " + str(e))

@task
def send_appointment_location_message(number, hospital_lat, hospital_long):
    data = {}
    data['phone_number'] = number

    link = '''http://maps.google.com/maps?q=loc:%s,%s''' % (hospital_lat, hospital_long)
    text = '''Location for your Upcoming Appointment %s ''' % (link)
    data['text'] = mark_safe(text)
    try:
        notification_models.SmsNotification.send_rating_link(data)
    except Exception as e:
        logger.error("Error sending Location message - " + str(e))

@task()
def process_payout(payout_id):
    from ondoc.account.models import MerchantPayout, Order, MerchantPayoutLog
    from ondoc.account.models import DummyTransactions
    from ondoc.doctor.models import OpdAppointment

    try:
        if not payout_id:
            raise Exception("No payout specified")

        payout_data = MerchantPayout.objects.filter(id=payout_id).first()
        if not payout_data:
            raise Exception("Payout not found")
        if payout_data.status == payout_data.PAID:
            raise Exception("Payment already done for this payout")

        default_payment_mode = payout_data.get_default_payment_mode()
        appointment = payout_data.get_appointment()
        if not appointment:
            raise Exception("Insufficient Data " + str(payout_data))

        if not payout_data.booking_type == payout_data.InsurancePremium:
            if appointment.payment_type in [OpdAppointment.COD]:
                raise Exception("Cannot process payout for COD appointments")

        billed_to = payout_data.get_billed_to()
        merchant = payout_data.get_merchant()
        order_data = None

        if not billed_to or not merchant:
            raise Exception("Insufficient Data " + str(payout_data))

        if not merchant.verified_by_finance or not merchant.enabled:
            raise Exception("Merchant is not verified or is not enabled. " + str(payout_data))

        if not payout_data.booking_type == payout_data.InsurancePremium:
            if payout_data.payout_type == 1:
                associated_merchant = billed_to.merchant.first()
                if not associated_merchant.verified:
                    raise Exception("Associated Merchant not verified. " + str(payout_data))

        # assuming 1 to 1 relation between Order and Appointment
        order_data = Order.objects.filter(reference_id=appointment.id).order_by('-id').first()

        if not order_data:
            raise Exception("Order not found for given payout " + str(payout_data))

        all_txn = order_data.getTransactions()

        if not all_txn or all_txn.count() == 0:
            raise Exception("No transactions found for given payout " + str(payout_data))

        nodal_id = payout_data.get_nodal_id
        req_data = {"payload": [], "checkSum": ""}
        req_data2 = {"payload": [], "checkSum": ""}

        idx = 0
        for txn in all_txn:
            
            curr_txn = OrderedDict()
            curr_txn["idx"] = idx
            curr_txn["orderNo"] = txn.order_no
            curr_txn["orderId"] = txn.order.id
            curr_txn["txnAmount"] = str(txn.amount)
            curr_txn["settledAmount"] = str(payout_data.payable_amount)
            curr_txn["merchantCode"] = merchant.id
            if txn.transaction_id and txn.transaction_id != 'null':
                curr_txn["pgtxId"] = txn.transaction_id
            curr_txn["refNo"] = payout_data.payout_ref_id
            curr_txn["bookingId"] = appointment.id
            curr_txn["paymentType"] = payout_data.payment_mode if payout_data.payment_mode else default_payment_mode
            curr_txn["nodalId"] = nodal_id
            req_data["payload"].append(curr_txn)
            idx += 1
            if isinstance(txn, DummyTransactions) and txn.amount>0:
                curr_txn2 = copy.deepcopy(curr_txn)
                curr_txn2["txnAmount"] = str(0)
                curr_txn2["idx"] = len(req_data2.get('payload'))
                req_data2["payload"].append(curr_txn2)
        payout_data.update_status('initiated')
        payout_status = None
        if len(req_data2.get('payload')) > 0:
            payout_status = request_payout(req_data2, order_data)
            payout_data.request_data = req_data2

        if not payout_status or not payout_status.get('status'):
            payout_status = request_payout(req_data, order_data)
            payout_data.request_data = req_data

        if payout_status:
            payout_data.api_response = payout_status.get("response")
            if payout_status.get("status"):
                payout_data.payout_time = datetime.datetime.now()
                # # Removed PAID status and add add initiated status when payout processed
                # payout_data.status = payout_data.PAID
                payout_data.status = payout_data.INPROCESS
            else:
                payout_data.retry_count += 1
                payout_data.status = payout_data.FAILED_FROM_QUEUE

            payout_data.save()

    except Exception as e:
        # update payout status
        payout_data.update_status('attempted')
        MerchantPayoutLog.create_log(payout_data, str(e))
        # logger.error("Error in processing payout - with exception - " + str(e))
        print("Error in processing payout - with exception - " + str(e))


@task(bind=True, max_retries=3)
def send_insurance_notifications(self, data):
    from ondoc.authentication import models as auth_model
    from ondoc.communications.models import InsuranceNotification
    from ondoc.insurance.models import UserInsurance
    try:
        user_id = int(data.get('user_id', 0))
        user = auth_model.User.objects.filter(id=user_id).last()
        if not user:
            raise Exception("Invalid user id passed for insurance email notification. Userid %s" % str(user_id))

        insurance_status = int(data.get('status', 0))
        # Cancellation
        if insurance_status and (insurance_status == UserInsurance.CANCEL_INITIATE
                                 or insurance_status == UserInsurance.CANCELLATION_APPROVED or
                                 insurance_status == UserInsurance.CANCELLED):
            user_insurance = UserInsurance.get_user_insurance(user)
            if not user_insurance:
                raise Exception("Invalid or None user insurance found for email notification. User id %s" % str(user_id))
            insurance_notification_status = None
            if insurance_status == UserInsurance.CANCEL_INITIATE:
                insurance_notification_status = NotificationAction.INSURANCE_CANCEL_INITIATE
            elif insurance_status == UserInsurance.CANCELLATION_APPROVED:
                insurance_notification_status = NotificationAction.INSURANCE_CANCELLATION_APPROVED
            elif insurance_status == UserInsurance.CANCELLED:
                insurance_notification_status = NotificationAction.INSURANCE_CANCELLATION
            insurance_notification = InsuranceNotification(user_insurance, insurance_notification_status)
            insurance_notification.send()

        else:
            user_insurance = user.active_insurance
            if not user_insurance:
                raise Exception("Invalid or None user insurance found for email notification. User id %s" % str(user_id))

            if not user_insurance.coi:
                try:
                    user_insurance.generate_pdf()
                except Exception as e:
                    logger.error('Insurance coi pdf cannot be generated. %s' % str(e))

                    countdown_time = (2 ** self.request.retries) * 60 * 10
                    print(countdown_time)
                    self.retry([data], countdown=countdown_time)

            insurance_notification = InsuranceNotification(user_insurance, NotificationAction.INSURANCE_CONFIRMED)
            insurance_notification.send()
    except Exception as e:
        logger.error(str(e))


@task(bind=True, max_retries=3)
def send_plus_membership_notifications(self, data):
    from ondoc.authentication import models as auth_model
    from ondoc.communications.models import VipNotification
    from ondoc.plus.models import PlusUser
    try:
        user_id = int(data.get('user_id', 0))
        user = auth_model.User.objects.filter(id=user_id).last()
        if not user:
            raise Exception("Invalid user id passed for plus membership email notification. Userid %s" % str(user_id))

        if user.active_plus_user:
            plus_user_obj = user.active_plus_user
        else:
            plus_user_obj = user.inactive_plus_user
        if not plus_user_obj:
            raise Exception("Invalid or None plus user membership found for email notification. User id %s" % str(user_id))

        plus_user_notification = VipNotification(plus_user_obj, NotificationAction.PLUS_MEMBERSHIP_CONFIRMED)
        plus_user_notification.send()
    except Exception as e:
        logger.error(str(e))


@task(bind=True, max_retries=3)
def send_insurance_endorsment_notifications(self, data):
    from ondoc.authentication import models as auth_model
    from ondoc.communications.models import InsuranceNotification
    from ondoc.insurance.models import UserInsurance, EndorsementRequest
    try:
        user_id = int(data.get('user_id', 0))
        user = auth_model.User.objects.filter(id=user_id).last()
        if not user:
            raise Exception("Invalid user id passed for insurance email notification. Userid %s" % str(user_id))

        user_insurance = user.active_insurance
        if not user_insurance:
            raise Exception("Invalid or None user insurance found for email notification. User id %s" % str(user_id))

        endorsment_status = data.get('endorsment_status', 0)
        notification = None

        if endorsment_status == EndorsementRequest.PENDING:
            notification = NotificationAction.INSURANCE_ENDORSMENT_PENDING
        elif endorsment_status == EndorsementRequest.REJECT:
            notification = NotificationAction.INSURANCE_ENDORSMENT_REJECTED
        elif endorsment_status == EndorsementRequest.APPROVED:
            notification = NotificationAction.INSURANCE_ENDORSMENT_APPROVED
        elif endorsment_status == EndorsementRequest.PARTIAL_APPROVED:
            notification = NotificationAction.INSURANCE_ENDORSMENT_PARTIAL_APPROVED

            if not user_insurance.coi:
                try:
                    user_insurance.generate_pdf()
                except Exception as e:
                    logger.error('Insurance coi pdf cannot be generated. %s' % str(e))

                    countdown_time = (2 ** self.request.retries) * 60 * 10
                    print(countdown_time)
                    self.retry([data], countdown=countdown_time)

        if notification and user_insurance:
            insurance_notification = InsuranceNotification(user_insurance, notification)
            insurance_notification.send()
            pending_members = user_insurance.endorse_members.filter(Q(mail_status=EndorsementRequest.MAIL_PENDING) |
                                                                    Q(mail_status__isnull=True))
            for member in pending_members:
                member.mail_status = EndorsementRequest.MAIL_SENT
                member.save()

    except Exception as e:
        logger.error(str(e))


@task(bind=True, max_retries=3)
def send_insurance_float_limit_notifications(self, data):
    from ondoc.notification.models import EmailNotification
    from ondoc.insurance.models import Insurer
    try:
        insurer_id = int(data.get('insurer_id', 0))
        insurer = Insurer.objects.filter(id=insurer_id).first()
        if not insurer:
            raise Exception('Insurer not found against the id %d' % insurer_id)

        insurer_account = insurer.float.filter().first()
        if not insurer_account:
            raise Exception('Insurer Account not found against the insurer id %d' % insurer_id)

        emails = settings.INSURANCE_FLOAT_LIMIT_ALERT_EMAIL
        html_body = "Insurer {insurer} current float amount is being getting exhausted and reached {limit}."\
            .format(insurer=insurer.name, limit=insurer_account.current_float)

        date = timezone.now() - timedelta(days=1)
        is_already_sent = EmailNotification.objects.filter(created_at__gte=date,
                                             notification_type=NotificationAction.INSURANCE_FLOAT_LIMIT).exists()

        if not is_already_sent:
            for email in emails:
                EmailNotification.send_insurance_float_alert_email(email, html_body)

    except Exception as e:
        logger.error(str(e))
        countdown_time = (2 ** self.request.retries) * 60 * 10
        print(countdown_time)
        self.retry([data], countdown=countdown_time)


def request_payout(req_data, order_data):
    from ondoc.api.v1.utils import create_payout_checksum
    from ondoc.account.mongo_models import PgLogs

    req_data["checkSum"] = create_payout_checksum(req_data["payload"], order_data.product_id)
    headers = {
        "auth": settings.PG_REFUND_AUTH_TOKEN,
        "Content-Type": "application/json"
    }
    url = settings.PG_SETTLEMENT_URL
    resp_data = None

    response = requests.post(url, data=json.dumps(req_data), headers=headers)
    resp_data = response.json()
    if settings.SAVE_LOGS:
        save_pg_response.apply_async((PgLogs.PAYOUT_PROCESS, order_data.id, None, resp_data, req_data, None), eta=timezone.localtime(), queue=settings.RABBITMQ_LOGS_QUEUE)
    if response.status_code == status.HTTP_200_OK:
        if resp_data.get("ok") is not None and resp_data.get("ok") == '1':
            success_payout = False
            result = resp_data.get('result')
            if result:
                for res_txn in result:
                    success_payout = res_txn['status'] == "SUCCESSFULLY_INSERTED"

            if success_payout:
                return {"status": 1, "response": resp_data}

    # logger.error("payout failed for request data - " + str(req_data))
    print("payout failed for request data - " + str(req_data))
    return {"status" : 0, "response" : resp_data}


@task()
def opd_send_otp_before_appointment(appointment_id, previous_appointment_date_time):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    try:
        instance = OpdAppointment.objects.filter(id=appointment_id).first()
        if not instance or \
                not instance.user or \
                str(math.floor(instance.time_slot_start.timestamp())) != previous_appointment_date_time \
                or instance.status != OpdAppointment.ACCEPTED:
            # logger.error(
            #     'instance : {}, time : {}, str: {}'.format(str(model_to_dict(instance)),
            #                                                previous_appointment_date_time,
            #                                                str(math.floor(instance.time_slot_start.timestamp()))))
            return
        opd_notification = OpdNotification(instance, NotificationAction.OPD_OTP_BEFORE_APPOINTMENT)
        opd_notification.send()
    except Exception as e:
        logger.error(str(e))

@task()
def docprime_appointment_reminder_sms_provider(appointment_id, appointment_updated_at):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    try:
        instance = OpdAppointment.objects.filter(id=appointment_id).first()
        if not instance or \
                not instance.user or \
                str(math.floor(instance.updated_at.timestamp())) != appointment_updated_at \
                or instance.status != OpdAppointment.ACCEPTED:
            # logger.error(
            #     'instance : {}, time : {}, str: {}'.format(str(model_to_dict(instance)),
            #                                                previous_appointment_date_time,
            #                                                str(math.floor(instance.time_slot_start.timestamp()))))
            return
        opd_notification = OpdNotification(instance, NotificationAction.DOCPRIME_APPOINTMENT_REMINDER_PROVIDER_SMS)
        opd_notification.send()
    except Exception as e:
        logger.error(str(e))


@task()
def offline_appointment_reminder_sms_patient(appointment_id, time_slot_start_timestamp, number):
    from ondoc.doctor.models import OfflineOPDAppointments
    from ondoc.communications.models import SMSNotification, OfflineOpdAppointments
    try:
        instance = OfflineOPDAppointments.objects.select_related('user', 'doctor', 'hospital')\
                                                 .prefetch_related('hospital__assoc_doctors')\
                                                 .filter(id=appointment_id).first()
        if not instance or \
                not instance.user or \
                math.floor(instance.time_slot_start.timestamp()) != time_slot_start_timestamp \
                or instance.status != OfflineOPDAppointments.ACCEPTED:
            # logger.error(
            #     'instance : {}, time : {}, str: {}'.format(str(model_to_dict(instance)),
            #                                                previous_appointment_date_time,
            #                                                str(math.floor(instance.time_slot_start.timestamp()))))
            return
        receivers = {'sms_receivers': [{"user": None, "phone_number": number}]}
        offline_opd_comm_obj = OfflineOpdAppointments(appointment=instance,
                                                      notification_type=NotificationAction.OFFLINE_APPOINTMENT_REMINDER_PROVIDER_SMS,
                                                      receivers=receivers)
        offline_opd_comm_obj.send()
        # data = {}
        # data['phone_number'] = kwargs.get('number')
        # data['text'] = mark_safe(kwargs.get('text'))
        # notification_models.SmsNotification.send_rating_link(data)
    except Exception as e:
        logger.error(str(e))


@task()
def opd_send_after_appointment_confirmation(appointment_id, previous_appointment_date_time, second=False):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    try:
        instance = OpdAppointment.objects.filter(id=appointment_id).first()
        if not instance or \
                not instance.user or \
                str(math.floor(instance.time_slot_start.timestamp())) != previous_appointment_date_time:
            return
        if instance.status == OpdAppointment.ACCEPTED and not instance.insurance:
            if not second:
                opd_notification = OpdNotification(instance, NotificationAction.OPD_CONFIRMATION_CHECK_AFTER_APPOINTMENT)
            else:
                opd_notification = OpdNotification(instance, NotificationAction.OPD_CONFIRMATION_SECOND_CHECK_AFTER_APPOINTMENT)
            opd_notification.send()
        if instance.status == OpdAppointment.COMPLETED and not instance.is_rated:
            opd_notification = OpdNotification(instance, NotificationAction.OPD_FEEDBACK_AFTER_APPOINTMENT)
            opd_notification.send()
    except Exception as e:
        logger.error(str(e))



@task()
def lab_send_after_appointment_confirmation(appointment_id, previous_appointment_date_time, second=False):
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.communications.models import LabNotification

    try:
        instance = LabAppointment.objects.filter(id=appointment_id).first()
        if not instance or \
                not instance.user or \
                str(math.floor(instance.time_slot_start.timestamp())) != previous_appointment_date_time:
            return
        if instance.status == LabAppointment.ACCEPTED:
            if not second:
                lab_notification = LabNotification(instance, NotificationAction.LAB_CONFIRMATION_CHECK_AFTER_APPOINTMENT)
            else:
                lab_notification = LabNotification(instance, NotificationAction.LAB_CONFIRMATION_SECOND_CHECK_AFTER_APPOINTMENT)
            lab_notification.send()
        if instance.status == LabAppointment.COMPLETED and not instance.is_rated:
            lab_notification = LabNotification(instance, NotificationAction.LAB_FEEDBACK_AFTER_APPOINTMENT)
            lab_notification.send()
    except Exception as e:
        logger.error(str(e))


@task()
def lab_send_otp_before_appointment(appointment_id, previous_appointment_date_time):
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.communications.models import LabNotification
    try:
        instance = LabAppointment.objects.filter(id=appointment_id).first()
        if not instance or \
                not instance.user or \
                str(math.floor(instance.time_slot_start.timestamp())) != previous_appointment_date_time \
                or instance.status != LabAppointment.ACCEPTED:
            # logger.error(
            #     'instance : {}, time : {}, str: {}'.format(str(model_to_dict(instance)),
            #                                                previous_appointment_date_time,
            #                                                str(math.floor(instance.time_slot_start.timestamp()))))
            return
        lab_notification = LabNotification(instance, NotificationAction.LAB_OTP_BEFORE_APPOINTMENT)
        lab_notification.send()
    except Exception as e:
        logger.error(str(e))


@task()
def send_lab_reports(appointment_id):
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.communications.models import LabNotification
    from ondoc.matrix.tasks import create_prescription_lead_to_matrix
    from ondoc.matrix.tasks import send_report_review_data_to_chat
    try:
        instance = LabAppointment.objects.filter(id=appointment_id).first()
        if not instance:
            return
        lab_notification = LabNotification(instance, NotificationAction.LAB_REPORT_SEND_VIA_CRM)
        is_valid_for_provider = True
        lab_notification.send(is_valid_for_provider)
        try:
            create_prescription_lead_to_matrix.apply_async(({'appointment_id': instance.id},), countdown=1)
        except Exception as e:
            logger.error(str(e))

        try:
            send_report_review_data_to_chat.apply_async(({'appointment_id': instance.id},), countdown=5)
        except Exception as e:
            logger.error(str(e))
            
    except Exception as e:
        logger.error(str(e))


@task()
def upload_doctor_data(obj_id):
    from ondoc.doctor.models import UploadDoctorData
    from ondoc.crm.management.commands import upload_doctor_data as upload_command
    instance = UploadDoctorData.objects.filter(id=obj_id).first()
    errors = []
    if not instance or not instance.status == UploadDoctorData.IN_PROGRESS:
        return
    try:
        source = instance.source
        batch = instance.batch
        lines = instance.lines if instance and instance.lines else 100000000
        wb = load_workbook(instance.file)
        sheets = wb.worksheets
        doctor = upload_command.UploadDoctor(errors)
        qualification = upload_command.UploadQualification(errors)
        experience = upload_command.UploadExperience(errors)
        membership = upload_command.UploadMembership(errors)
        award = upload_command.UploadAward(errors)
        hospital = upload_command.UploadHospital(errors)
        specialization = upload_command.UploadSpecialization(errors)
        with transaction.atomic():
            # doctor.p_image(sheets[0], source, batch)
            doctor.upload(sheets[0], source, batch, lines, instance.user)
            qualification.upload(sheets[1], lines)
            experience.upload(sheets[2], lines)
            membership.upload(sheets[3], lines)
            award.upload(sheets[4], lines)
            hospital.upload(sheets[5], source, batch, lines)
            specialization.upload(sheets[6], lines)
            if len(errors)>0:
                raise Exception('errors in data')
        instance.status = UploadDoctorData.SUCCESS
        instance.save()
    except Exception as e:
        error_message = traceback.format_exc() + str(e)
        logger.error(error_message)
        instance.status = UploadDoctorData.FAIL
        if errors:
            instance.error_msg = errors
        else:
            instance.error_msg = [{'line number': 0, 'message': error_message}]
        instance.save(retry=False)

@task()
def send_pg_acknowledge(order_id=None, order_no=None, ack_type=''):
    from ondoc.account.mongo_models import PgLogs
    log_requests_on()
    try:
        if order_id is None or order_no is None:
            logger.error("Cannot acknowledge without order_id and order_no")
            return

        url = settings.PG_PAYMENT_ACKNOWLEDGE_URL + "?orderNo=" + str(order_no) + "&orderId=" + str(order_id)
        if ack_type == 'capture':
            url += "&ack=captureAck"
        response = requests.get(url)
        if response.status_code == status.HTTP_200_OK:
            if ack_type == 'capture':
                print("Payment capture acknowledged")
            else:
                print("Payment acknowledged")
        json_url = '{"url": "%s"}' % url
        if settings.SAVE_LOGS:
            save_pg_response.apply_async((PgLogs.ACK_TO_PG, order_id, None, None, json_url, None), eta=timezone.localtime(), queue=settings.RABBITMQ_LOGS_QUEUE)
    except Exception as e:
        logger.error("Error in sending pg acknowledge - " + str(e))




@task()
def refund_completed_sms_task(obj_id):
    from ondoc.account.models import ConsumerRefund
    from ondoc.communications.models import SMSNotification
    from ondoc.notification.models import NotificationAction
    try:
        # check if any more obj incomplete status
        instance = ConsumerRefund.objects.filter(id=obj_id).first()
        if not instance or not instance.user or ConsumerRefund.objects.filter(
                consumer_transaction_id=instance.consumer_transaction_id,
                refund_state__in=[ConsumerRefund.PENDING, ConsumerRefund.REQUESTED]).count() > 1:
            return

        context = {'amount': instance.consumer_transaction.amount, 'ctrnx_id': instance.consumer_transaction_id}
        receivers = instance.user.get_phone_number_for_communication()
        sms_notification = SMSNotification(NotificationAction.REFUND_COMPLETED, context)
        sms_notification.send(receivers)
    except Exception as e:
        logger.error(str(e))


@task()
def refund_breakup_sms_task(obj_id):
    from ondoc.account.models import ConsumerTransaction
    from ondoc.communications.models import SMSNotification
    try:
        instance = ConsumerTransaction.objects.filter(id=obj_id).first()
        if not instance or not instance.user:
            return
        context = {'amount': instance.amount, 'ctrnx_id': instance.id}
        receivers = instance.user.get_phone_number_for_communication()
        sms_notification = SMSNotification(NotificationAction.REFUND_BREAKUP, context)
        sms_notification.send(receivers)
    except Exception as e:
        logger.error(str(e))


@task(bind=True, max_retries=2)
def push_plus_lead_to_matrix(self, data):
    from ondoc.plus.models import PlusLead, PlusPlans
    from ondoc.notification.tasks import save_matrix_logs
    from ondoc.plus.models import PlusPlanUtmSources
    try:
        if not data:
            raise Exception('Data not received for banner lead.')

        id = data.get('id', None)
        if not id:
            logger.error("[CELERY ERROR: Incorrect values provided.]")
            raise ValueError()

        plus_lead_obj = PlusLead.objects.filter(id=id).first()
        obj_type = 'plus_lead'

        if not plus_lead_obj:
            raise Exception("Plus lead object could not found against id - " + str(id))

        if plus_lead_obj.user:
            phone_number = plus_lead_obj.user.phone_number
        else:
            phone_number = plus_lead_obj.phone_number

        extras = plus_lead_obj.extras
        plan_id = extras.get('plan_id', None)
        lead_data = extras.get('lead_data', {})

        lead_utm_source = lead_data.get('utm_source')
        if lead_utm_source and PlusPlanUtmSources.objects.filter(source=lead_utm_source, create_lead=False).exists():
            return False

        lead_source = "Docprime"

        if lead_data:
            provided_lead_source = lead_data.get('lead_source')
            # if provided_lead_source:
            #     lead_source = provided_lead_source
            if type(provided_lead_source).__name__ == 'str' and provided_lead_source.lower() == 'AppointmentPaySuccess'.lower():
                lead_source = provided_lead_source  #TODO change

        plan = None
        if plan_id and type(plan_id).__name__ == 'int':
            plan = PlusPlans.objects.filter(id=plan_id).first()

        request_data = {
            # 'LeadID': plus_lead_obj.matrix_lead_id if plus_lead_obj.matrix_lead_id else 0,
            'LeadID':  0,
            'LeadSource': lead_source,
            'Name': extras.get('name', 'none') if extras.get('name', 'none') else "none",
            'BookedBy': phone_number,
            'PrimaryNo': phone_number,
            'PaymentStatus': 0,
            'UtmCampaign': extras.get('utm_campaign', ''),
            'UTMMedium': extras.get('utm_medium', ''),
            'UtmSource': extras.get('utm_source', ''),
            'UtmTerm': extras.get('utm_term', ''),
            'ProductId': 11,
            'SubProductId': 0,
            'VIPPlanName': plan.plan_name if plan else None,
            'IsInsured': 1 if plus_lead_obj and plus_lead_obj.user and plus_lead_obj.user.active_insurance else 0
        }

        url = settings.MATRIX_API_URL
        matrix_api_token = settings.MATRIX_API_TOKEN
        response = requests.post(url, data=json.dumps(request_data), headers={'Authorization': matrix_api_token,
                                                                              'Content-Type': 'application/json'})

        # MatrixLog.create_matrix_logs(plus_lead_obj, request_data, response.json())
        if settings.SAVE_LOGS:
            save_matrix_logs.apply_async((plus_lead_obj.id, obj_type, request_data, response.json()), countdown=5, queue=settings.RABBITMQ_LOGS_QUEUE)

        if response.status_code != status.HTTP_200_OK or not response.ok:
            logger.error(json.dumps(request_data))
            logger.info("[ERROR] Plus lead could not be published to the matrix system")
            logger.info("[ERROR] %s", response.reason)

            countdown_time = (2 ** self.request.retries) * 60 * 10
            logging.error("Lead sync with the Matrix System failed with response - " + str(response.content))
            print(countdown_time)
            self.retry([data], countdown=countdown_time)
        else:
            resp_data = response.json()
            if not resp_data:
                raise Exception('Data received from matrix is null or empty.')

            if not resp_data.get('Id', None):
                return
                # logger.error(json.dumps(request_data))
                # raise Exception("[ERROR] Id not recieved from the matrix while pushing plus lead to matrix.")

            plus_lead_qs = PlusLead.objects.filter(id=id)
            plus_lead_qs.update(matrix_lead_id=resp_data.get('Id'))

    except Exception as e:
        logger.error("Error in Celery. Failed pushing Plus lead to the matrix- " + str(e))


@task(bind=True, max_retries=2)
def push_insurance_banner_lead_to_matrix(self, data):
    from ondoc.insurance.models import InsuranceLead, InsurancePlans
    from ondoc.notification.tasks import save_matrix_logs
    try:
        if not data:
            raise Exception('Data not received for banner lead.')

        id = data.get('id', None)
        if not id:
            logger.error("[CELERY ERROR: Incorrect values provided.]")
            raise ValueError()

        banner_obj = InsuranceLead.objects.filter(id=id).first()
        obj_type = 'insurance_lead'

        if not banner_obj:
            raise Exception("Banner object could not found against id - " + str(id))

        if banner_obj.user:
            phone_number = banner_obj.user.phone_number
        else:
            phone_number = banner_obj.phone_number

        extras = banner_obj.extras
        plan_id = extras.get('plan_id', None)

        lead_source = "InsuranceOPD"
        lead_data = extras.get('lead_data')
        if lead_data:
            provided_lead_source = lead_data.get('source')
            if type(provided_lead_source).__name__ == 'str' and provided_lead_source.lower() == 'docprimechat':
                lead_source = 'docprimechat'

        plan = None
        if plan_id and type(plan_id).__name__ == 'int':
            plan = InsurancePlans.objects.filter(id=plan_id).first()

        request_data = {
            'LeadID': banner_obj.matrix_lead_id if banner_obj.matrix_lead_id else 0,
            'LeadSource': lead_source,
            'Name': 'none',
            'BookedBy': phone_number,
            'PrimaryNo': phone_number,
            'PaymentStatus': 0,
            'UtmCampaign': extras.get('utm_campaign', ''),
            'UTMMedium': extras.get('utm_medium', ''),
            'UtmSource': extras.get('utm_source', ''),
            'UtmTerm': extras.get('utm_term', ''),
            'ProductId': 8,
            'SubProductId': 0,
            'PolicyDetails': {
                "ProposalNo": None,
                "BookingId": None,
                'PolicyPaymentSTATUS': 0,
                "ProposerName": None,
                "PolicyId": None,
                "InsurancePlanPurchased": plan.name if plan else None,
                "PurchaseDate": None,
                "ExpirationDate": None,
                "COILink": None,
                "PeopleCovered": 0
            }
        }

        url = settings.MATRIX_API_URL
        matrix_api_token = settings.MATRIX_API_TOKEN
        response = requests.post(url, data=json.dumps(request_data), headers={'Authorization': matrix_api_token,
                                                                              'Content-Type': 'application/json'})

        # MatrixLog.create_matrix_logs(banner_obj, request_data, response.json())
        if settings.SAVE_LOGS:
            save_matrix_logs.apply_async((banner_obj.id, obj_type, request_data, response.json()), countdown=5, queue=settings.RABBITMQ_LOGS_QUEUE)

        if response.status_code != status.HTTP_200_OK or not response.ok:
            logger.error(json.dumps(request_data))
            logger.info("[ERROR] Insurance banner lead could not be published to the matrix system")
            logger.info("[ERROR] %s", response.reason)

            countdown_time = (2 ** self.request.retries) * 60 * 10
            logging.error("Lead sync with the Matrix System failed with response - " + str(response.content))
            print(countdown_time)
            self.retry([data], countdown=countdown_time)
        else:
            resp_data = response.json()
            if not resp_data:
                raise Exception('Data received from matrix is null or empty.')

            if not resp_data.get('Id', None):
                return
                # logger.error(json.dumps(request_data))
                # raise Exception("[ERROR] Id not recieved from the matrix while pushing insurance banner lead to matrix.")

            insurance_banner_qs = InsuranceLead.objects.filter(id=id)
            insurance_banner_qs.update(matrix_lead_id=resp_data.get('Id'))

    except Exception as e:
        logger.error("Error in Celery. Failed pushing insurance banner lead to the matrix- " + str(e))


@task
def generate_random_coupons(total_count, coupon_id):
    from ondoc.coupon.models import RandomGeneratedCoupon, Coupon
    try:
        coupon_obj = Coupon.objects.filter(id=coupon_id).first()
        if not coupon_obj:
            return

        while total_count:
            curr_count = 0
            batch_data = []
            while curr_count < 10000 and total_count:
                rc = RandomGeneratedCoupon()
                rc.random_coupon = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(12))
                rc.coupon = coupon_obj
                rc.validity = 90
                rc.sent_at = datetime.datetime.utcnow()

                batch_data.append(rc)
                curr_count += 1
                total_count -= 1

            if batch_data:
                RandomGeneratedCoupon.objects.bulk_create(batch_data)
            else:
                return

    except Exception as e:
        logger.error(str(e))

@task
def update_random_coupons_consumption(random_coupons_ids=None):
    from ondoc.coupon.models import RandomGeneratedCoupon
    if random_coupons_ids:
        random_coupons = RandomGeneratedCoupon.objects.filter(random_coupon=random_coupons_ids)
        random_coupons.update(consumed_at=datetime.datetime.now())


@task
def update_coupon_used_count():
    RawSql('''  update coupon set total_used_count= usage_count from
                (select coupon_id, sum(usage_count) usage_count from
                (select oac.coupon_id, count(*) usage_count from opd_appointment oa inner join opd_appointment_coupon oac on oa.id = oac.opdappointment_id
                where oa.status in (2,3,4,5,7) group by oac.coupon_id
                union
                select lac.coupon_id, count(*) usage_count from lab_appointment la inner join lab_appointment_coupon lac on la.id = lac.labappointment_id
                where la.status in (2,3,4,5,7) group by lac.coupon_id
                union
                select puc.coupon_id, count(*) usage_count from plus_users pu inner join plus_users_coupon puc on pu.id = puc.plususer_id
                where pu.status in (1,3,4,5,6,7) group by puc.coupon_id
                ) x group by coupon_id
                ) y where coupon.id = y.coupon_id ''', []).execute()


@task
def send_ipd_procedure_cost_estimate(ipd_procedure_lead_id=None):
    from ondoc.procedure.models import IpdProcedureLead
    from ondoc.communications.models import IpdLeadNotification
    try:
        instance = IpdProcedureLead.objects.filter(id=ipd_procedure_lead_id).first()
        ipd_lead_notification = IpdLeadNotification(ipd_procedure_lead=instance, notification_type=NotificationAction.IPD_PROCEDURE_COST_ESTIMATE)
        ipd_lead_notification.send()
    except Exception as e:
        logger.error(str(e))


@task()
def upload_cost_estimates(obj_id):
    from ondoc.procedure.models import UploadCostEstimateData
    from ondoc.crm.management.commands import upload_cost_estimates as upload_command
    instance = UploadCostEstimateData.objects.filter(id=obj_id).first()
    errors = []
    if not instance or not instance.status == UploadCostEstimateData.IN_PROGRESS:
        return
    try:
        wb = load_workbook(instance.file)
        sheets = wb.worksheets
        cost_estimate = upload_command.UploadCostEstimate(errors)
        cost_estimate.upload(sheets[0])
        if len(errors)>0:
            raise Exception('errors in data')
        instance.status = UploadCostEstimateData.SUCCESS
        instance.save()
    except Exception as e:
        error_message = traceback.format_exc() + str(e)
        logger.error(error_message)
        instance.status = UploadCostEstimateData.FAIL
        if errors:
            instance.error_msg = errors
        else:
            instance.error_msg = [{'line number': 0, 'message': error_message}]
        instance.save(retry=False)


@task()
def send_user_number_update_otp(obj_id):
    from ondoc.sms.backends.backend import BaseSmsBackend

    obj = UserNumberUpdate.objects.filter(id=obj_id).first()
    if not obj:
        return

    phone_numer = obj.new_number
    otp = obj.otp

    sms_class = BaseSmsBackend()
    message = "Otp for new number update :%s . Dont share this with anyone." % str(otp)
    success = sms_class.send(message, phone_numer)
    if not success:
        logger.error("Could not send otp for user number update.")

    return


@task()
def send_userprofile_email_update_otp(obj_id):
    from ondoc.notification.models import EmailNotification
    obj = UserProfileEmailUpdate.objects.filter(id=obj_id).first()
    if not obj:
        return

    EmailNotification.send_userprofile_email_update(obj)

    return


@task()
def send_contactus_notification(obj_id):
    from ondoc.notification.models import EmailNotification
    from ondoc.web.models import ContactUs
    from django.contrib.contenttypes.models import ContentType
    obj = ContactUs.objects.filter(id=obj_id).first()

    if not obj:
        return

    emails = settings.CONTACTUS_EMAILS

    html_body = "{name} ( {email}-{mobile} ) has sent message {message}" \
        .format(name=obj.name, email=obj.email, mobile=obj.mobile, message={obj.message})

    mobile_number = None
    if obj.mobile:
        mobile_number = obj.mobile

    if obj.from_app:
        html_body += " from mobile app."
    else:
        html_body += "."

    content_type = ContentType.objects.get_for_model(obj)
    date = timezone.now() - timedelta(days=1)
    is_already_sent = EmailNotification.objects.filter(created_at__gte=date,
                                                       notification_type=NotificationAction.CONTACT_US_EMAIL,
                                                       content_type=content_type,
                                                       object_id=obj.id).exists()

    if not is_already_sent:
        for email in emails:
            EmailNotification.send_contact_us_notification_email(content_type, obj.id, email, html_body, mobile_number)


@task(bind=True, max_retries=5)
def send_capture_payment_request(self, product_id, appointment_id):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.account.models import Order, PgTransaction, PaymentProcessStatus
    from ondoc.account.mongo_models import PgLogs
    from ondoc.account.models import ConsumerTransaction
    log_requests_on()
    req_data = dict()
    if product_id == Order.DOCTOR_PRODUCT_ID:
        obj = OpdAppointment
    if product_id == Order.LAB_PRODUCT_ID:
        obj = LabAppointment
    try:
        order = Order.objects.filter(product_id=product_id, reference_id=appointment_id).first()

        appointment = order.getAppointment()
        if not appointment:
            raise Exception("No Appointment found.")

        if order and not order.is_parent():
            order = order.parent

        txn_obj = PgTransaction.objects.filter(order=order).first()

        # if not appointment.status == obj.CANCELLED:
        if txn_obj and txn_obj.is_preauth():
            url = settings.PG_CAPTURE_PAYMENT_URL
            token = settings.PG_SEAMLESS_CAPTURE_AUTH_TOKEN

            req_data = {
                "orderNo": txn_obj.order_no,
                "orderId": str(order.id),
                "hash": pg_seamless_hash(order, txn_obj.order_no)
            }

            headers = {
                "auth": token,
                "Content-Type": "application/json"
            }

            response = requests.post(url, data=json.dumps(req_data), headers=headers)

            resp_data = response.json()
            if settings.SAVE_LOGS:
                save_pg_response.apply_async((PgLogs.TXN_CAPTURED, order.id, txn_obj.id, resp_data, req_data, order.user_id,), eta=timezone.localtime(), queue=settings.RABBITMQ_LOGS_QUEUE)

            args = {'order_id': order.id, 'status_code': resp_data.get('statusCode'), 'source': 'CAPTURE'}
            status_type = PaymentProcessStatus.get_status_type(resp_data.get('statusCode'), resp_data.get('txStatus'))
            save_payment_status.apply_async((status_type, args), eta=timezone.localtime(), )
            if response.status_code == status.HTTP_200_OK:
                txn_captured = False
                if resp_data.get("ok") is not None and resp_data.get("ok") == '1':
                    txn_obj.status_code = resp_data.get('statusCode')
                    txn_obj.status_type = resp_data.get('txStatus')
                    txn_obj.payment_mode = resp_data.get("paymentMode")
                    txn_obj.bank_name = resp_data.get('bankName')
                    txn_obj.transaction_id = resp_data.get('pgTxId')
                    txn_obj.bank_id = resp_data.get('bankTxId')
                    txn_obj.payment_captured = True
                    ctx_txn = ConsumerTransaction.objects.filter(order_id=order.id, action=ConsumerTransaction.PAYMENT).last()
                    ctx_txn.transaction_id = resp_data.get('pgTxId')
                    ctx_txn.save()
                    txn_captured = True
                else:
                    txn_obj.payment_captured = False
                    logger.error("Error in capture the payment with data - " + json.dumps(req_data) + " with error message - " + resp_data.get('statusMsg', ''))
                txn_obj.save()

                if txn_captured:
                    send_pg_acknowledge.apply_async((txn_obj.order_id, txn_obj.order_no, 'capture'), countdown=1)
            else:
                raise Exception("Retry on invalid Http response status - " + str(response.content))

    except Exception as e:
        logger.error("Error in payment capture with data - " + json.dumps(req_data) + " with exception - " + str(e))
        self.retry([product_id, appointment_id], countdown=300)


@task(bind=True, max_retries=5)
def send_release_payment_request(self, product_id, appointment_id):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.account.models import Order, PgTransaction, PaymentProcessStatus
    from ondoc.account.mongo_models import PgLogs
    log_requests_on()
    req_data = dict()
    if product_id == Order.DOCTOR_PRODUCT_ID:
        obj = OpdAppointment
    if product_id == Order.LAB_PRODUCT_ID:
        obj = LabAppointment
    try:
        order = Order.objects.filter(product_id=product_id, reference_id=appointment_id).first()

        appointment = order.getAppointment()
        if not appointment:
            raise Exception("No Appointment found.")

        if order and not order.is_parent():
            order = order.parent

        txn_obj = PgTransaction.objects.filter(order=order).first()

        # if appointment.status == obj.CANCELLED:
        if txn_obj and txn_obj.is_preauth():

            tz = pytz.timezone(settings.TIME_ZONE)
            if datetime.datetime.now(tz) < (txn_obj.transaction_date + datetime.timedelta(
                    hours=int(settings.PAYMENT_AUTO_CAPTURE_DURATION))).astimezone(tz):
                url = settings.PG_RELEASE_PAYMENT_URL
                token = settings.PG_SEAMLESS_RELEASE_AUTH_TOKEN

                req_data = {
                    "orderNo": txn_obj.order_no,
                    "orderId": str(order.id),
                    "hash": pg_seamless_hash(order, txn_obj.order_no)
                }

                headers = {
                    "auth": token,
                    "Content-Type": "application/json"
                }

                response = requests.post(url, data=json.dumps(req_data), headers=headers)
                resp_data = response.json()
                if settings.SAVE_LOGS:
                    save_pg_response.apply_async((PgLogs.TXN_RELEASED, order.id, txn_obj.id, resp_data, req_data, order.user_id,), eta=timezone.localtime(), queue=settings.RABBITMQ_LOGS_QUEUE)

                args = {'order_id': order.id, 'status_code': resp_data.get('statusCode'), 'source': 'RELEASE'}
                status_type = PaymentProcessStatus.get_status_type(resp_data.get('statusCode'),
                                                                   resp_data.get('txStatus'))
                save_payment_status.apply_async((status_type, args), eta=timezone.localtime(), )
                if response.status_code == status.HTTP_200_OK:
                    if resp_data.get("ok") is not None and resp_data.get("ok") == '1':
                        txn_obj.status_code = resp_data.get('statusCode')
                        txn_obj.status_type = 'TXN_RELEASE'
                        txn_obj.save()
                        send_pg_acknowledge.apply_async((txn_obj.order_id, txn_obj.order_no, 'capture'), countdown=1)
                    else:
                        logger.error("Error in releasing the payment with data - " + json.dumps(
                            req_data) + " with error message - " + resp_data.get('statusMsg', ''))
                else:
                    raise Exception("Retry on invalid Http response status - " + str(response.content))

    except Exception as e:
        logger.error("Error in payment release with data - " + json.dumps(req_data) + " with exception - " + str(e))
        self.retry([product_id, appointment_id], countdown=300)

@task(bind=True)
def save_pg_response(self, log_type, order_id, txn_id, response, request, user_id, log_created_at=None, *args, **kwargs):
    try:
        from ondoc.account.mongo_models import PgLogs
        if order_id.__class__.__name__ == 'list':
            PgLogs.save_single_pg_response(log_type, order_id, txn_id, response, request, user_id)
        else:
            if response:
                if not isinstance(response, dict):
                    response = json.loads(response)
                response.pop('created_at', None)
            PgLogs.save_pg_response(log_type, order_id, txn_id, response, request, user_id, log_created_at)
    except Exception as e:
        # todo - temporary commented to avoid error logs in sentry
        # logger.error("Error in saving pg response to mongo database - " + json.dumps(response) + " with exception - " + str(e))
        pass


@task(bind=True)
def save_payment_status(self, current_status, args):
    try:
        from ondoc.account.models import PaymentProcessStatus
        if args.get('is_single'):
            PaymentProcessStatus.save_single_payment_status(current_status, args)
        else:
            PaymentProcessStatus.save_payment_status(current_status, args)
    except Exception as e:
       logger.error("Error in saving payment status - " + json.dumps(args) + " with exception - " + str(e))


@task()
def purchase_order_creation_counter_automation(purchase_order_id):

    from ondoc.doctor.models import PurchaseOrderCreation
    instance = PurchaseOrderCreation.objects.filter(id=purchase_order_id).first()
    if instance:
        if (instance.start_date < instance.end_date):
            instance.is_enabled = True
            instance.provider_name_hospital.enabled_poc = True
            instance.provider_name_hospital.enabled_for_cod = True

            # TODO: In OPDAppointment
            # if timezone.now() >= instance.end_date:
            #     instance.is_enabled = False
            #     instance.provider_name_hospital.enabled_poc = False
            #     instance.provider_name_hospital.enabled_for_cod = False



@task()
def purchase_order_closing_counter_automation(purchase_order_id):

    from ondoc.doctor.models import PurchaseOrderCreation
    instance = PurchaseOrderCreation.objects.filter(id=purchase_order_id).first()
    if instance:
        if (instance.end_date < timezone.now().date()):
            instance.disable_cod_functionality()


@task(bind=True)
def send_lensfit_coupons(self, appointment_id, product_id, notification_type=None):
    from ondoc.account.models import Order
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.coupon.models import Coupon
    from ondoc.communications.models import OpdNotification
    from ondoc.communications.models import LabNotification
    try:
        if product_id == Order.DOCTOR_PRODUCT_ID:
            cls = OpdAppointment
        elif product_id == Order.LAB_PRODUCT_ID:
            cls = LabAppointment
        appointment_obj =cls.objects.filter(pk=appointment_id).first()
        if not appointment_obj or not appointment_obj.user:
            return
        if appointment_obj.status == OpdAppointment.COMPLETED:
            lensfit_coupon = Coupon.get_lensfit_coupon()
            if product_id == Order.DOCTOR_PRODUCT_ID:
                notification = OpdNotification(appointment_obj, notification_type, {"lensfit_coupon": lensfit_coupon})
            elif product_id == Order.LAB_PRODUCT_ID:
                notification = LabNotification(appointment_obj, notification_type, {"lensfit_coupon": lensfit_coupon})
            notification.send()
    except Exception as e:
        logger.error(str(e))


@task()
def send_partner_lab_notifications(order_id, notification_type=None, report_list=list()):
    from ondoc.provider.models import PartnerLabSamplesCollectOrder
    from ondoc.communications.models import PartnerLabNotification
    try:
        if not order_id:
            return
        instance = PartnerLabSamplesCollectOrder.objects.select_related('doctor', 'hospital', 'offline_patient', 'created_by') \
                                                        .prefetch_related('lab_alerts', 'reports') \
                                                        .filter(id=order_id).first()
        if not instance:
            return
        comm_kwargs = dict()
        if notification_type:
            comm_kwargs['notification_type'] = notification_type
            comm_kwargs['report_list'] = report_list
        partner_lab_comm_obj = PartnerLabNotification(instance, **comm_kwargs)
        partner_lab_comm_obj.send()
    except Exception as e:
        logger.error(str(e))


@task(bind=True, max_retries=3)
def save_matrix_logs(self, id, obj_type, request_data, response):
    try:
        from ondoc.matrix.mongo_models import MatrixLog
        from ondoc.diagnostic.models import LabAppointment
        from ondoc.doctor.models import OpdAppointment
        from ondoc.insurance.models import UserInsurance, InsuranceLead
        from ondoc.plus.models import PlusUser, PlusLead
        from ondoc.common.models import GeneralMatrixLeads

        object = None
        if obj_type == 'lab_appointment':
            object = LabAppointment.objects.filter(id=id).first()
        elif obj_type == 'opd_appointment':
            object = OpdAppointment.objects.filter(id=id).first()
        elif obj_type == 'user_insurance':
            object = UserInsurance.objects.filter(id=id).first()
        elif obj_type == 'plus_user':
            object = PlusUser.objects.filter(id=id).first()
        elif obj_type == 'plus_lead':
            object = PlusLead.objects.filter(id=id).first()
        elif obj_type == 'insurance_lead':
            object = InsuranceLead.objects.filter(id=id).first()
        elif obj_type == 'general_leads':
            object = GeneralMatrixLeads.objects.filter(id=id).first()

        MatrixLog.create_matrix_logs(object, request_data, response)

    except Exception as e:
        logger.error("Error in saving matrix logs to mongo database - " + json.dumps(response) + " with exception - " + str(e))


@task(bind=True, max_retries=2)
def process_leads_to_matrix(self, data):
    from ondoc.common.models import GeneralMatrixLeads
    from ondoc.common.lead_engine import lead_class_referance
    from ondoc.plus.models import PlusUser
    from ondoc.doctor.models import OpdAppointment
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.insurance.models import UserInsurance

    try:
        if not data:
            raise Exception('Data not received for general matrix leads.')

        id = data.get('id', None)
        if not id:
            logger.error("[CELERY ERROR: Incorrect values provided.]")
            raise ValueError()

        general_lead_obj = GeneralMatrixLeads.objects.filter(id=id).first()

        if not general_lead_obj:
            raise Exception("GeneralMatrix object could not found against id - " + str(id))

        if general_lead_obj.lead_type in ['DROPOFF', 'LABADS', 'CANCELDROPOFFLEADVIAAPPOINTMENT', 'DOCADS']:
            plus_obj = PlusUser.objects.filter(user__phone_number=general_lead_obj.phone_number).order_by('-id').first()
            insurance_obj = UserInsurance.objects.filter(user__phone_number=general_lead_obj.phone_number).order_by('-id').first()
            earlier_date = timezone.now() - timedelta(minutes=10)
            lab_appointment = LabAppointment.objects.filter(created_at__gte=earlier_date, user__phone_number=general_lead_obj.phone_number).first()
            opd_appointment = OpdAppointment.objects.filter(created_at__gte=earlier_date, user__phone_number=general_lead_obj.phone_number).first()

            if (plus_obj and plus_obj.is_valid()) or (insurance_obj and insurance_obj.is_valid()):
                return

            if general_lead_obj.lead_type == 'DROPOFF' and (lab_appointment or opd_appointment):
                return

        lead_engine_obj = lead_class_referance(general_lead_obj.lead_type, general_lead_obj)
        success = lead_engine_obj.process_lead()

        if not success:
            countdown_time = (2 ** self.request.retries) * 60 * 10
            print(countdown_time)
            self.retry([data], countdown=countdown_time)

    except Exception as e:
        logger.error("Error in Celery. Failed pushing General lead to the matrix- " + str(e))


@task()
def opd_send_completion_notification(appointment_id, payment_type):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    try:
        instance = OpdAppointment.objects.filter(id=appointment_id).first()
        if not instance or not instance.user:
            return
        if payment_type in (OpdAppointment.PREPAID, OpdAppointment.INSURANCE, OpdAppointment.VIP, OpdAppointment.GOLD, OpdAppointment.PLAN):
            opd_notification = OpdNotification(instance, NotificationAction.PROVIDER_OPD_APPOINTMENT_COMPLETION_ONLINE_PAYMENT)
            opd_notification.send()
        if payment_type == OpdAppointment.COD:
            opd_notification = OpdNotification(instance, NotificationAction.PROVIDER_OPD_APPOINTMENT_COMPLETION_PAY_AT_CLINIC)
            opd_notification.send()
    except Exception as e:
        logger.error(str(e))

@task()
def opd_send_confirmation_notification(data):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    try:
        instance = OpdAppointment.objects.filter(id=data.get('appointment_id')).first()
        if not instance or not instance.user:
            return

        if data.get('payment_type') in (OpdAppointment.PREPAID, OpdAppointment.INSURANCE, OpdAppointment.VIP, OpdAppointment.GOLD, OpdAppointment.PLAN):
            opd_notification = OpdNotification(instance, NotificationAction.PROVIDER_OPD_APPOINTMENT_CONFIRMATION_ONLINE_PAYMENT)
            opd_notification.send()
        if data.get('payment_type') == OpdAppointment.COD:
            opd_notification = OpdNotification(instance, NotificationAction.PROVIDER_OPD_APPOINTMENT_CONFIRMATION_PAY_AT_CLINIC)
            opd_notification.send()
    except Exception as e:
        logger.error(str(e))

@task()
def lab_send_notification_before_appointment(appointment_id, time):
    from ondoc.diagnostic.models import LabAppointment
    from ondoc.communications.models import LabNotification
    try:
        instance = LabAppointment.objects.filter(id=appointment_id).first()
        if not instance or not instance.user:
            return
        lab_notification = LabNotification(instance, NotificationAction.PROVIDER_LAB_APPOINTMENT_CONFIRMATION_ONLINE_PAYMENT)
        lab_notification.send()
    except Exception as e:
        logger.error(str(e))


@task()
def push_reminder_message_medanta_and_artemis(data, *args, **kwargs):
    from ondoc.doctor.models import OpdAppointment
    from ondoc.communications.models import OpdNotification
    try:
        instance = OpdAppointment.objects.filter(id=data.get('appointment_id')).first()
        if not instance or not instance.user:
            return
        opd_notification = OpdNotification(instance, NotificationAction.REMINDER_MESSAGE_MEDANTA_AND_ARTEMIS)
        receivers = opd_notification.get_receivers(is_valid_for_provider=True)

        from ondoc.communications.models import SMSNotification
        context = {'doctor_name': instance.doctor.name.title(),
                   'time_slot_start_date': instance.time_slot_start.date(),
                   'time_slot_start_time': instance.time_slot_start.strftime("%I:%M%p"),
                   'hospital_name': instance.hospital.name,
                   'patient_name': instance.profile_detail.get('name').title() if instance.profile_detail.get(
                       'name') else ''}
        sms_notification = SMSNotification(NotificationAction.REMINDER_MESSAGE_MEDANTA_AND_ARTEMIS, context)
        sms_notification.send(receivers.get('sms_receivers', []), *args, **kwargs)

    except Exception as e:
        logger.error(str(e))