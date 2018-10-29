from django.core.management.base import BaseCommand
from django.conf import settings
from io import BytesIO
from openpyxl import load_workbook
import requests
import re
from PIL import Image as Img
import os
import math
from django.core.files.storage import default_storage
from ondoc.doctor.models import (Doctor, DoctorPracticeSpecialization, PracticeSpecialization, DoctorMobile,
                                 Qualification,
                                 Specialization, College, DoctorQualification, DoctorExperience, DoctorAward,
                                 DoctorClinicTiming, DoctorClinic, Hospital, SourceIdentifier, DoctorAssociation,
                                 DoctorPopularity)
from django.contrib.gis.geos import Point, GEOSGeometry


class Command(BaseCommand):
    help = 'Upload doctors via Excel'

    def add_arguments(self, parser):
        parser.add_argument('url', type=str, help='data url')

    def handle(self, *args, **options):
        print(options)
        url = options['url']
        r = requests.get(url)
        content = BytesIO(r.content)
        wb = load_workbook(content)
        sheets = wb.worksheets
        doctor = UploadDoctor()

        doctor.upload(sheets[0])


class UploadDoctor:

    popularity_mapping = {value[1]: value[0] for value in DoctorPopularity.POPULARITY_CHOICES}

    def upload(self, sheet):
        rows = [row for row in sheet.rows]
        headers = {column.value.strip().lower(): i + 1 for i, column in enumerate(rows[0]) if column.value}
        all_data = []
        for i in range(2, len(rows) + 1):
            data = self.get_data(row=i, sheet=sheet, headers=headers)
            all_data.append(DoctorPopularity(**data))
        DoctorPopularity.objects.bulk_create(all_data)

    def get_data(self, row, sheet, headers):
        unique_identifier = self.clean_data(sheet.cell(row=row, column=headers.get('unique_identifier')).value)
        popularity = self.popularity_mapping.get(self.clean_data(sheet.cell(row=row, column=headers.get('popularity')).value))
        popularity_score = self.clean_data(sheet.cell(row=row, column=headers.get('popularity_score')).value)
        rating_percent = self.clean_data(sheet.cell(row=row, column=headers.get('rating_percent')).value)
        votes_count = self.clean_data(sheet.cell(row=row, column=headers.get('votes_count')).value)
        reviews_count = self.clean_data(sheet.cell(row=row, column=headers.get('reviews_count')).value)
        try:
            popularity = int(popularity)
        except:
            print('Invalid popularity ='+str(popularity))
            popularity = DoctorPopularity.NON_KEY

        try:
            popularity_score = float(popularity_score)
        except:
            print('Invalid popularity_score='+str(popularity_score))
            popularity_score = 0

        try:
            rating_percent = int(rating_percent)
        except:
            print('Invalid rating_percent='+str(rating_percent))
            rating_percent = 0

        try:
            votes_count = int(votes_count)
        except:
            print('Invalid votes_count='+str(votes_count))
            votes_count = 0

        try:
            reviews_count = int(reviews_count)
        except:
            print('Invalid reviews_count='+str(reviews_count))
            reviews_count = 0

        data = {'unique_identifier': unique_identifier, 'popularity': popularity, 'popularity_score': popularity_score,
                'rating_percent': rating_percent, 'votes_count': votes_count, 'reviews_count': reviews_count}
        return data

    def create_doctor(self, data, source, batch):
        return

    @staticmethod
    def clean_data(value):
        if value and isinstance(value, str):
            return value.strip()
        return value
